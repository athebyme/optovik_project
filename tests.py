from __future__ import print_function

import io

import google.auth
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload

# print(input())
import openpyxl
import os

# xlsx = openpyxl.load_workbook(os.getcwd()+'\\report_2022_7_29.xlsx')
# sheet = xlsx.active
# col_names= []
#
# for row in sheet.iter_rows(1,sheet.max_row):
#     print(row[3].value)


# set_of_post_codes = [
#     'Z1C1A0t-', 'W1C1Ayt', 'Z1C1L', 'Z1C1K', 'W1C1K', 'W1C1V', 'Z1C1A', 'W1C1L', 'W1C1A'
# ]
# item = '1234Z1C1L9645'
# if any(set_of_post_codes) not in item:
#     pass
# else:
#     print(set_of_post_codes.index(item))

# from pyfiglet import Figlet
#
# @staticmethod
# def showText(text):
#     try:
#         view_text = Figlet(font='slant')
#         print(view_text.renderText(text))
#     except:
#         print(text)
#
#
# showText('DNIS')


#
# def download_file(real_file_id):
#     """Downloads a file
#     Args:
#         real_file_id: ID of the file to download
#     Returns : IO object with location.
#
#     Load pre-authorized user credentials from the environment.
#     TODO(developer) - See https://developers.google.com/identity
#     for guides on implementing OAuth2 for the application.
#     """
#     creds, _ = google.auth.default()
#
#     try:
#         # create drive api client
#         service = build('drive', 'v3', credentials=creds)
#
#         file_id = real_file_id
#
#         # pylint: disable=maybe-no-member
#         request = service.files().get_media(fileId=file_id)
#         file = io.BytesIO()
#         downloader = MediaIoBaseDownload(file, request)
#         done = False
#         while done is False:
#             status, done = downloader.next_chunk()
#             print(F'Download {int(status.progress() * 100)}.')
#
#     except HttpError as error:
#         print(F'An error occurred: {error}')
#         file = None
#
#     return file.getvalue()
#
#
# if __name__ == '__main__':
#     download_file(real_file_id='1iOAtmIzmDJZZ1vRkWC0tOiyod6uP5ZugRI1hoJhqXjc')


# from itertools import product
#
# @staticmethod
# def init_category(data):
#     CONST_COUNT_WORDS = 3
#     data_lower = data.lower()
#     getting_in_cat_wb = set()
#     cats_wb = set()
#     info = list(map(lambda data: data.strip(), data_lower.lower().split('#')))
#     info = list(map(lambda info: info.split('>'), info))
#     with open('E:\csv_parser_wb\pool\SexOptovik\google_downloaded\wb_cats.txt') as f:
#         for i in f:
#             cats_wb.add(i.rstrip().lower())
#     cats_wb.remove('')
#
#     for i in info:
#         saved = []
#         for j in i:
#             if 'белье' not in j:
#                 for n in j.split():
#                     if n in cats_wb:
#                         getting_in_cat_wb.add(n)
#                 t = list(map(lambda j: j+' ', j.split()))
#                 for p in saved:
#                     t.append(p+' ')
#                 for n in range(CONST_COUNT_WORDS):
#                     t.append('')
#                 if 'эротик' not in t:
#                     t.append('эротик ')
#                 all_var = set(map(''.join,product(t,repeat = CONST_COUNT_WORDS)))
#                 all_var = list(map(lambda i:i.rstrip(),all_var))
#                 all_var = list(map(lambda i: i.replace(',',''),all_var))
#                 for var in all_var:
#                     if var in cats_wb:
#                         getting_in_cat_wb.add(var)
#                 saved = j.split().copy()
#             else:
#                 t = data_lower.lower()
#                 if 'комплекты' in t:
#                     getting_in_cat_wb.add('комплекты эротик')
#                 if 'бдсм' in t:
#                     getting_in_cat_wb.add('комплекты бдсм')
#                 if 'топы' in t:
#                     getting_in_cat_wb.add('топы эротик')
#                 if 'лифы' in t:
#                     getting_in_cat_wb.add('бюстгальтеры эротик')
#                 if 'бюстье' in t:
#                     getting_in_cat_wb.add('бюстгальтеры эротик')
#                 if 'стрэпы' in t:
#                     getting_in_cat_wb.add('стрэпы эротик')
#                 if 'корсаж' in t:
#                     getting_in_cat_wb.add('корсеты эротик')
#                 if 'трусики' in t or 'трусы' in t or 'стринги' in t or 'шортики' in t or 'шорты' in t:
#                     getting_in_cat_wb.add('трусы эротик')
#                 if 'маски' in t:
#                     getting_in_cat_wb.add('маски эротик')
#     if 'вибро' in data_lower:
#         if 'зажим' in data_lower:
#             getting_in_cat_wb.add('зажимы для сосков')
#         if 'яйца' in data_lower:
#             getting_in_cat_wb.add('виброяйца')
#         if 'трус' in data_lower:
#             getting_in_cat_wb.add('вибротрусики')
#         if 'пуля' or 'пули' in data_lower:
#             getting_in_cat_wb.add('вибропули')
#     ed = ['лубрикан', 'стек ', 'смазки для секс-игрушек', 'наборы и аксессуары', 'клиторальные стимуляторы',
#            'костюмы', 'бдсм товары и фетиш > аксессуары', 'наручник', 'кольцо эрекционное', 'стимулирующие влагалище',
#            ' танго', 'стреп']
#     cats_ed = ['лубриканты', 'стэки эротик', 'уходовые средства эротик', 'наборы игрушек для взрослых', 'вибраторы',
#                'ролевые костюмы эротик', 'комплекты бдсм', 'наручник эротик', 'эрекционные кольца',
#                'вагинальные тренажеры', 'трусы эротик','стрэпы эротик']
#     for o in range(len(ed)):
#         if ed[o] in data_lower:
#             getting_in_cat_wb.add(cats_ed[o])
#     res = []
#     if len(getting_in_cat_wb)>1 and 'комплекты бдсм' in getting_in_cat_wb:
#         getting_in_cat_wb.remove('комплекты бдсм')
#     for i in getting_in_cat_wb:
#         cat = i[:1].upper() + i[1:]
#         if 'БДСМ' in i:
#             cat = cat.replace('бдсм','БДСМ')
#         res.append(cat)
#     if len(res) == 0:
#         res.append('Наборы игрушек для взрослых')
#     return res[0], data.replace(' #','.')
#
# print(init_category('Красный блесиящий стреп H.E.L. Essie'))

# print(os.path.exists('./pool/SexOptovik/google_downloaded/wb_cats2.txt'))

#
#
# country = 'Китай-США' # Страна производства
# if '-' in country:
#     country = country[:country.find('-')]
# print(country)
# import re
#
# @staticmethod
# def for_sizes_parse(array_4XL):
#     size = ['XS','S','M','L','XL','XXL','XXXL','4XL']
#     size_ru = array_4XL.copy()
#     dict = {}
#     for i in range(len(size)):
#         dict.setdefault(size[i],size_ru[i])
#     return dict
# def parse_sizes(text,type):
#     data = list(map(str,text.split(', ')))
#     res = {}
#     if type == 'all':
#         for item in data:
#             it = item
#             if ' мм' in it:
#                 k = 10
#             elif ' кг' in it:
#                 k = 0.001
#             else:
#                 k = 1
#
#             #if '-' in it:
#             #    it = it[it.find('-')+1:]
#             if 'вес' in it and res.setdefault('weight') is None:
#                 w = 0
#                 if text.lower().count('вес') > 1:
#                     for i in data:
#                         if 'вес' in i:
#                             w += round(float(re.sub("[^0-9,]", "", i))/k,2)
#                 else:
#                     w = round(int(re.sub("[^0-9,]", "", it))/k,2)
#                 res['weight'] = w
#
#             eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
#             eds_en = ['length', 'width', 'height', 'diameter', 'depth']
#             for i in range(len(eds)):
#                 if eds[i] in it and res.setdefault(eds_en[i]) is None:
#                     _r = re.sub("[^0-9,-]", "", it).replace(',','.')
#                     if '-' in r:
#                         r_lst = list(map(str,r.split('-')))
#                         av = 0
#                         for j in r_lst:
#                             av += float(j)
#                         _r = round(float(av/len(r_lst)) / k, 2)
#                     else:
#                         _r = round(float(_r) / k, 2)
#                     res[eds_en[i]] = _r
#                     break
#         return res
#     if type == 'clothe':
#         SIZES = {1:'XS',2:'S',3:'M',4:'L',5:'XL',6:'XXL',7:'XXXL',8:'4XL'}
#         SIZES_RU = {'XS':'40-42','S':'42-44','M':'44-46','L':'46-48','XL':'48-50','1XL':'52-54','XXL':'54-56',
#                     'XXXL':'56-58','4XL':'58-60','XL':'50-52'}
#         for k,v in SIZES_RU.items():
#             if v in text:
#                 res['size'] = k
#                 res['size-ru'] = v
#         check = list(map(str,text.split(', ')))
#         for i in check:
#             if 'об. ' in i and res.setdefault('size') is None:
#                 n = re.sub("[^0-9-]", "", i)
#                 l = list(map(int, n.split('-')))
#                 sr_arf = round((l[1] + l[0]) / 2)
#                 if 'груди' in i:
#                     SIZE = for_sizes_parse([99,103,107,111,115,119,123,127])
#                 if 'бедер' in i:
#                     SIZE = for_sizes_parse([103,107,111,115,119,123,127,131])
#                 if 'талия' in i:
#                     SIZE = for_sizes_parse([86,90,94,98,102,107,112,117])
#                 if sr_arf < SIZE['XS']:
#                     res['size'] = 'XS'
#                     res['size-ru'] = SIZES_RU['S']
#                     return res
#                 prev_sz = SIZE['XS']
#                 for k,v in SIZE.items():
#                     if k == 'XS': pass
#                     if sr_arf < v and sr_arf > prev_sz:
#                         res['size'] = k
#                         res['size-ru'] = SIZES_RU[k]
#                         return res
#                     prev_sz = v
#
#         t = check[0]
#         if t.count('.') > 1:
#             res['size'] = 'M'
#             res['size-ru'] = SIZES_RU['M']
#             return res
#         if 'универсальный' in t:
#             res['size'] = 'M'
#             res['size-ru'] = SIZES_RU['M']
#             return res
#         if '-' in t and len(t) == 5:
#             res['size-ru'] = t
#             size = t.split('-')[0]
#             for k,v in SIZES_RU.items():
#                 if size in v:
#                     res['size'] = k
#             return res
#         for k,v in SIZES.items():
#             if (v in t or str(k) in t) and res.setdefault('size') is None:
#                 res['size'] = v
#                 res['size-ru'] = SIZES_RU[v]
#                 return res
#         return res
#     if type == 'cosmetic':
#         ed = ''
#         if ' мл' in text:
#             ed = 'мл'
#         if ' г' in text:
#             ed = 'г'
#         volume = re.sub("[^0-9 *хx,]", "", text)
#         volume = volume.replace('  ', ' *', 1)
#         res.setdefault('volume', volume.strip())
#         res.setdefault('ed', ed.strip())
#         return res
#
# size = parse_sizes('общая длина 23,5 см, диаметр 5,7-9 см',type='all')
# r = []
# r.append(size.get('123'))
# print(size.get('depth'), r)

# СОХРАНИТЬ
# @staticmethod
# def init_category(data, cats_wb):
#     CONST_COUNT_WORDS = 3
#     data_lower = data.lower()
#     getting_in_cat_wb = set()
#     cats_wb = set()
#     info = list(map(lambda data: data.strip(), data_lower.lower().split('#')))
#     info = list(map(lambda info: info.split('>'), info))
#     with open('E:\csv_parser_wb\pool\SexOptovik\google_downloaded\wb_cats.txt') as f:
#         for i in f:
#             cats_wb.add(i.rstrip().lower())
#     cats_wb.remove('')
#     if 'вибро' in data_lower:
#         if 'зажим' in data_lower:
#             getting_in_cat_wb.add('зажимы для сосков')
#         if 'яйца' in data_lower:
#             getting_in_cat_wb.add('виброяйца')
#         if 'трус' in data_lower:
#             getting_in_cat_wb.add('вибротрусики')
#         if 'пуля' or 'пули' in data_lower:
#             getting_in_cat_wb.add('вибропули')
#     if 'смазки для секс-игрушек' in data_lower:
#         getting_in_cat_wb.add('уходовые средства эротик')
#     if 'наборы и аксессуары' in data_lower:
#         getting_in_cat_wb.add('наборы игрушек для взрослых')
#     if 'клиторальные стимуляторы' in data_lower:
#         getting_in_cat_wb.add('вибраторы')
#     if 'костюмы' in data_lower:
#         getting_in_cat_wb.add('ролевые костюмы эротик')
#     if 'бдсм товары и фетиш > аксессуары' in data_lower:
#         getting_in_cat_wb.add('комплекты бдсм')
#     if 'наручник' in data_lower:
#         getting_in_cat_wb.add('наручник эротик')
#     if 'презерватив' in data_lower:
#         getting_in_cat_wb.add('презервативы')
#     for i in info:
#         saved = []
#         for j in i:
#             if 'белье' not in j:
#                 for n in j.split():
#                     if n in cats_wb:
#                         getting_in_cat_wb.add(n)
#                 t = list(map(lambda j: j + ' ', j.split()))
#                 for p in saved:
#                     t.append(p + ' ')
#                 for n in range(CONST_COUNT_WORDS):
#                     t.append('')
#                 # t.append('эротик')
#                 all_var = list(map(''.join, product(t, repeat=CONST_COUNT_WORDS)))
#                 all_var = list(map(lambda i: i.rstrip(), all_var))
#                 all_var = list(map(lambda i: i.replace(',', ''), all_var))
#                 for var in all_var:
#                     if var in cats_wb:
#                         getting_in_cat_wb.add(var)
#                 saved = j.split().copy()
#             else:
#                 t = data_lower.lower()
#                 if 'комплекты' in t:
#                     getting_in_cat_wb.add('комплекты эротик')
#                 if 'бдсм' in t:
#                     getting_in_cat_wb.add('комплекты бдсм')
#                 if 'топы' in t:
#                     getting_in_cat_wb.add('топы эротик')
#                 if 'лифы' in t:
#                     getting_in_cat_wb.add('бюстгальтеры эротик')
#                 if 'бюстье' in t:
#                     getting_in_cat_wb.add('бюстгальтеры эротик')
#                 if 'стрэпы' in t:
#                     getting_in_cat_wb.add('стрэпы эротик')
#                 if 'корсаж' in t:
#                     getting_in_cat_wb.add('корсеты эротик')
#                 if 'трусики' in t or 'трусы' in t or 'стринги' in t or 'шортики' in t or 'шорты' in t:
#                     getting_in_cat_wb.add('трусы эротик')
#                 if 'маски' in t:
#                     getting_in_cat_wb.add('маски эротик')
#     res = []
#     if len(getting_in_cat_wb) > 1 and 'комплекты бдсм' in getting_in_cat_wb:
#         getting_in_cat_wb.remove('комплекты бдсм')
#     for i in getting_in_cat_wb:
#         cat = i[:1].upper() + i[1:]
#         if 'БДСМ' in i:
#             cat = cat.replace('бдсм', 'БДСМ')
#         res.append(cat)
#     if len(res) == 0:
#         res.append('Наборы игрушек для взрослых')
#     return res[0], data.replace(' #', '.')

# def start(text):
#     if ', ' not in text:
#         _split_parametr = '  '
#         eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
#         indx = []
#         data = []
#         for x in eds:
#             if text.find(x) != -1: indx.append(text.find(x))
#         indx.sort()
#         if len(indx) > 1:
#             for x in range(1,len(indx)):
#                 data.append(text[:indx[x]-1].strip())
#                 text = text[indx[x]:]
#             data.append(text.strip())
#         else:
#             data.append(text)
#         return data
# print(start('длина 17,75 см,диаметр 11,5 см'))


# import re
# @staticmethod
# def for_sizes_parse(array_4XL):
#     size = ['XS', 'S', 'M', 'L', 'XL', 'XXL', 'XXXL', '4XL']
#     size_ru = array_4XL.copy()
#     dict = {}
#     for i in range(len(size)):
#         dict.setdefault(size[i], size_ru[i])
#     return dict
# @staticmethod
# def parse_sizes(text, type):
#     t = text
#     data = []
#     eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина', 'вес']
#     indx = []
#     indx_first = []
#     data = []
#     for x in eds:
#         if t.find(x) != -1: indx.append(t.find(x))
#     indx_first = indx.copy()
#     if len(indx) > 0:
#         if indx[0] != 0:
#             t = t[indx[0]:]
#             indx = []
#             for x in eds:
#                 if t.find(x) != -1: indx.append(t.find(x))
#         indx.sort()
#     if len(indx) != 0:
#         if len(indx) > 1:
#             for x in range(1, len(indx)):
#                 data.append(t[:indx[x] - 1].strip())
#                 t = t[indx[x]:]
#             data.append(t.strip())
#         else:
#             if text.strip() not in data:
#                 data.append(text.strip())
#     res = {}
#     if type == 'all':
#         for item in data:
#             it = item
#             if ' мм' in it:
#                 k = 10
#             elif ' кг' in it:
#                 k = 0.001
#             else:
#                 k = 1
#             # it = it.replace(',', '.')
#             # if '-' in it:
#             #    it = it[it.find('-')+1:]
#             if 'вес' in it and res.setdefault('weight') is None:
#                 it = it.replace(',', '.')
#                 w = 0
#                 if text.lower().count('вес') > 1:
#                     for i in data:
#                         if 'вес' in i:
#                             w += round(float(re.sub("[^0-9,]", "", i).replace(',','.')) / k, 2)
#                 else:
#                     w = round(float(re.sub("[^0-9,]", "", it).replace(',','.')) / k, 2)
#                 res['weight'] = w
#
#             eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
#             eds_en = ['length', 'width', 'height', 'diameter', 'depth']
#             for i in range(len(eds)):
#                 if eds[i] in it and res.setdefault(eds_en[i]) is None:
#                     r = re.sub("[^0-9,-]", "", it).replace(',', '.')
#                     if r == '':
#                         res[eds_en[i]] = 'универсальный (растягивается)'
#                         break
#                     if '-' in r:
#                         r_lst = list(map(str, r.split('-')))
#                         try:
#                             r_lst.remove('')
#                         except ValueError:
#                             pass
#                         av = 0
#                         for j in r_lst:
#                             av += float(j)
#                         r = round(float(av / len(r_lst)) / k, 2)
#                     else:
#                         r = round(float(r) / k, 2)
#                     res[eds_en[i]] = r
#                     break
#         return res
#     if type == 'clothe':
#         SIZES = {1: 'XS', 2: 'S', 3: 'M', 4: 'L', 5: 'XL', 6: 'XXL', 7: 'XXXL', 8: '4XL'}
#         SIZES_RU = {'XS': '40-42', 'S': '42-44', 'M': '44-46', 'L': '46-48', 'XL': '48-50', '1XL': '52-54',
#                     'XXL': '54-56',
#                     'XXXL': '56-58', '4XL': '58-60', 'XL': '50-52'}
#         eds = ['об. талии', 'об. груди', 'об. бедер']
#         indx_obhvats = []
#         for i in eds:
#             if i in text:
#                 indx_obhvats.append(text.index(i))
#         if len(data) != 0:
#             eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
#             eds_en = ['length', 'width', 'height', 'diameter', 'depth']
#             for x in data:
#                 temp = x
#                 if ' мм' in temp:
#                     k = 10
#                 elif ' кг' in temp:
#                     k = 0.001
#                 else:
#                     k = 1
#                 for z in range(len(eds)):
#                     if eds[z] in x:
#                         t = re.sub("[^0-9,-]", "", x).replace(',', '.')
#                         if res.setdefault(eds_en[z]) is None:
#                             res[eds_en[z]] = t
#                 if 'вес' in temp and res.setdefault('weight') is None:
#                     it = temp.replace(',', '.')
#                     w = 0
#                     if text.lower().count('вес') > 1:
#                         for i in data:
#                             if 'вес' in i:
#                                 w += round(float(re.sub("[^0-9,]", "", i)) / k, 2)
#                     else:
#                         w = round(float(re.sub("[^0-9,]", "", it)) / k, 2)
#                     res['weight'] = w
#
#         for k, v in SIZES_RU.items():
#             if v in text:
#                 res['size'] = k
#                 res['size-ru'] = v
#         if len(indx_obhvats) != 0 and len(indx_first) != 0 :
#             if indx_first[len(indx_first) - 1] > indx_obhvats[len(indx_obhvats) - 1]:
#                 text = text[:indx_first[0]]
#             else:
#                 pass
#         check = list(map(str, text.split(', ')))
#
#         for i in check:
#             if 'об. ' in i and res.setdefault('size') is None:
#                 n = re.sub("[^0-9-]", "", i)
#                 l = list(map(int, n.split('-')))
#                 sr_arf = round((l[1] + l[0]) / 2)
#                 if 'груд' in i:
#                     SIZE = for_sizes_parse([99, 103, 107, 111, 115, 119, 123, 127])
#                 if 'бедер' in i:
#                     SIZE = for_sizes_parse([103, 107, 111, 115, 119, 123, 127, 131])
#                 if 'тали' in i:
#                     SIZE = for_sizes_parse([86, 90, 94, 98, 102, 107, 112, 117])
#                 if sr_arf < SIZE['XS']:
#                     res['size'] = 'XS'
#                     res['size-ru'] = SIZES_RU['S']
#                     return res
#                 prev_sz = SIZE['XS']
#                 for k, v in SIZE.items():
#                     if k == 'XS': pass
#                     if sr_arf < v and sr_arf > prev_sz:
#                         res['size'] = k
#                         res['size-ru'] = SIZES_RU[k]
#                         return res
#                     prev_sz = v
#
#         t = check[0]
#         if t.count('.') > 1:
#             res['size'] = 'M'
#             res['size-ru'] = SIZES_RU['M']
#             return res
#         if 'универсальный' in t:
#             res['size'] = 'M'
#             res['size-ru'] = SIZES_RU['M']
#             return res
#         if '-' in t and len(t) == 5:
#             res['size-ru'] = t
#             size = t.split('-')[0]
#             for k, v in SIZES_RU.items():
#                 if size in v:
#                     res['size'] = k
#             return res
#         for k, v in SIZES.items():
#             if (v in t or str(k) in t) and res.setdefault('size') is None:
#                 res['size'] = v
#                 res['size-ru'] = SIZES_RU[v]
#                 return res
#         return res
#     if type == 'cosmetic':
#         ed = ''
#         if ' мл' in text:
#             ed = 'мл'
#         if ' г' in text:
#             ed = 'г'
#         volume = re.sub("[^0-9 *хx,]", "", text)
#         volume = volume.replace('  ', ' *', 1)
#         res.setdefault('volume', volume.strip())
#         res.setdefault('ed', ed.strip())
#         return res
#     return text
# print(parse_sizes('диаметр шариков  - 2,8 см, вес белого - 14 г, вес серого - 29 г, вес темно-серого 36 г, вес черного 55 г',type = 'all'))


import re


@staticmethod
def for_sizes_parse(array_4XL):
    size = ['XS', 'S', 'M', 'L', 'XL', 'XXL', 'XXXL', '4XL']
    size_ru = array_4XL.copy()
    dict = {}
    for i in range(len(size)):
        dict.setdefault(size[i], size_ru[i])
    return dict

# a = ''
# a = float(10)
# print(a)

import pathlib as Path
# path = './shrih/Список штрихкодов (4).xlsx'
# #with ope n
# with open(path,encoding='ansi') as f:
#     for lines in f:
#         print(lines)


# @staticmethod
# def parse_sizes(text, type):
#     eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
#     eds_en = ['length', 'width', 'height', 'diameter', 'depth']
#     if ', ' not in text or type == 'clothe':
#         t = text
#         data = []
#         eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина', 'вес']
#         indx = []
#         indx_first = []
#         data = []
#         for x in eds:
#             if t.find(x) != -1: indx.append(t.find(x))
#         indx.sort()
#         indx_first = indx.copy()
#         if len(indx) != 0:
#             if indx[0] != 0:
#                 t = t[indx[0]:]
#                 indx = []
#                 for x in eds:
#                     if t.find(x) != -1: indx.append(t.find(x))
#
#             if len(indx) > 1:
#                 for x in range(1, len(indx)):
#                     data.append(t[:indx[x] - 1].strip())
#                     t = t[indx[x]:]
#                 data.append(t.strip())
#             else:
#                 if text.strip() not in data:
#                     data.append(text.strip())
#     else:
#         _split_parametr = ', '
#         data = list(map(str, text.split(_split_parametr)))
#     res = {}
#     ed_local_check = [',длина', ',ширина', ',высота', ',диаметр', ',глубина']
#     for i in data:
#         FLAG = []
#         for b in ed_local_check:
#             if b in i:
#                 FLAG.append(True)
#         if len(FLAG) > 1:
#             _indx = []
#             divided = []
#             for j in range(len(ed_local_check)):
#                 if i.find(eds[j]) != -1: _indx.append(i.find(eds[j]))
#             _indx.sort()
#             data.remove(i)
#             t = i
#             for x in range(1, len(_indx)):
#                 data.append(t[:_indx[x] - 1].strip())
#                 t = t[_indx[x]:]
#             data.append(t.strip())
#
#     ed_local_check = [',длина', ',ширина', ',высота', ',диаметр', ',глубина']
#     for i in data:
#         FLAG = []
#         for b in ed_local_check:
#             if b in i:
#                 FLAG.append(True)
#         if len(FLAG) >= 1:
#             _indx = []
#             divided = []
#             for j in range(len(ed_local_check)):
#                 if i.find(eds[j]) != -1: _indx.append(i.find(eds[j]))
#             _indx.sort()
#             data.remove(i)
#             t = i
#             for x in range(1, len(_indx)):
#                 data.append(t[:_indx[x] - 1].strip())
#                 t = t[_indx[x]:]
#             data.append(t.strip())
#     if type == 'all':
#         for item in data:
#             it = item
#             if ' мм' in it:
#                 k = 10
#             elif ' кг' in it:
#                 k = 0.001
#             else:
#                 k = 1
#             # it = it.replace(',', '.')
#             # if '-' in it:
#             #    it = it[it.find('-')+1:]
#             if 'вес' in it and res.setdefault('weight') is None:
#                 it = it.replace(',', '.')
#                 w = 0
#                 if text.lower().count('вес') > 1:
#                     for i in data:
#                         if 'вес' in i:
#                             w += round(float(re.sub("[^0-9,]", "", i).replace(',', '.')) / k, 2)
#                 else:
#                     w = round(float(re.sub("[^0-9,]", "", it).replace(',', '.')) / k, 2)
#                 res['weight'] = w
#
#             for i in range(len(eds)):
#                 if eds[i] in it and res.setdefault(eds_en[i]) is None:
#                     r = re.sub("[^0-9,-]", "", it).replace(',', '.')
#                     if r == '':
#                         res[eds_en[i]] = 'универсальный (растягивается)'
#                         break
#                     if '-' in r:
#                         r_lst = list(map(str, r.split('-')))
#                         try:
#                             r_lst.remove('')
#                         except ValueError:
#                             pass
#                         av = 0
#                         for j in r_lst:
#                             av += float(j)
#                         r = round(float(av / len(r_lst)) / k, 2)
#                     else:
#                         r = round(float(r) / k, 2)
#                     res[eds_en[i]] = r
#                     break
#         return res
#     if type == 'clothe':
#         SIZES = {1: 'XS', 2: 'S', 3: 'M', 4: 'L', 5: 'XL', 6: 'XXL', 7: 'XXXL', 8: '4XL'}
#         SIZES_RU = {'XS': '40-42', 'S': '42-44', 'M': '44-46', 'L': '46-48', 'XL': '48-50', '1XL': '52-54',
#                     'XXL': '54-56',
#                     'XXXL': '56-58', '4XL': '58-60', 'XL': '50-52'}
#         eds = ['об. талии', 'об. груди', 'об. бедер']
#         indx_obhvats = []
#         for i in eds:
#             if i in text:
#                 indx_obhvats.append(text.index(i))
#         if len(data) != 0:
#             eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
#             eds_en = ['length', 'width', 'height', 'diameter', 'depth']
#             for x in data:
#                 temp = x
#                 if ' мм' in temp:
#                     k = 10
#                 elif ' кг' in temp:
#                     k = 0.001
#                 else:
#                     k = 1
#                 for z in range(len(eds)):
#                     if eds[z] in x:
#                         t = re.sub("[^0-9,-]", "", x).replace(',', '.')
#                         if res.setdefault(eds_en[z]) is None:
#                             res[eds_en[z]] = t
#                 if 'вес' in temp and res.setdefault('weight') is None:
#                     it = temp.replace(',', '.')
#                     w = 0
#                     if text.lower().count('вес') > 1:
#                         for i in data:
#                             if 'вес' in i:
#                                 w += round(float(re.sub("[^0-9,]", "", i)) / k, 2)
#                     else:
#                         w = round(float(re.sub("[^0-9,]", "", it)) / k, 2)
#                     res['weight'] = w
#
#         for k, v in SIZES_RU.items():
#             if v in text:
#                 res['size'] = k
#                 res['size-ru'] = v
#         if len(indx_obhvats) != 0 and len(indx_first) != 0:
#             if indx_first[len(indx_first) - 1] > indx_obhvats[len(indx_obhvats) - 1]:
#                 text = text[:indx_first[0]]
#             else:
#                 pass
#         check = list(map(str, text.split(', ')))
#
#         for i in check:
#             if 'об. ' in i and res.setdefault('size') is None:
#                 n = re.sub("[^0-9-]", "", i)
#                 l = list(map(int, n.split('-')))
#                 sr_arf = round((l[1] + l[0]) / 2)
#                 if 'груд' in i:
#                     SIZE = for_sizes_parse([99, 103, 107, 111, 115, 119, 123, 127])
#                 if 'бедер' in i:
#                     SIZE = for_sizes_parse([103, 107, 111, 115, 119, 123, 127, 131])
#                 if 'тали' in i:
#                     SIZE = for_sizes_parse([86, 90, 94, 98, 102, 107, 112, 117])
#                 if sr_arf < SIZE['XS']:
#                     res['size'] = 'XS'
#                     res['size-ru'] = SIZES_RU['S']
#                     return res
#                 prev_sz = SIZE['XS']
#                 for k, v in SIZE.items():
#                     if k == 'XS': pass
#                     if sr_arf < v and sr_arf > prev_sz:
#                         res['size'] = k
#                         res['size-ru'] = SIZES_RU[k]
#                         return res
#                     prev_sz = v
#
#         t = check[0]
#         if t.count('.') > 1:
#             res['size'] = 'M'
#             res['size-ru'] = SIZES_RU['M']
#             return res
#         if 'универсальный' in t:
#             res['size'] = 'M'
#             res['size-ru'] = SIZES_RU['M']
#             return res
#         if '-' in t and len(t) == 5:
#             res['size-ru'] = t
#             size = t.split('-')[0]
#             for k, v in SIZES_RU.items():
#                 if size in v:
#                     res['size'] = k
#             return res
#         for k, v in SIZES.items():
#             if (v in t or str(k) in t) and res.setdefault('size') is None:
#                 res['size'] = v
#                 res['size-ru'] = SIZES_RU[v]
#                 return res
#         return res
#     if type == 'cosmetic':
#         ed = ''
#         if ' мл' in text:
#             ed = 'мл'
#         if ' г' in text:
#             ed = 'г'
#         volume = re.sub("[^0-9 *хx,]", "", text)
#         volume = volume.replace('  ', ' *', 1)
#         res.setdefault('volume', volume.strip())
#         res.setdefault('ed', ed.strip())
#         return res
#     return text
# print(parse_sizes(
#     'общая длина бутылки 28 см, макс. диаметр 7,2 см,глубина проникновения 14 см, внутренний диаметр 1,5 см (растягивается)',
#     type='all'))

# from itertools import product
#
# #t =
# CONST_COUNT_WORDS = 3
# all_var = set(map(''.join, product(t, repeat=CONST_COUNT_WORDS)))
# all_var = list(map(lambda i: i.rstrip().replace(',', ''), all_var))
# a = 'наручники эротик'
# b = 'наруч эротик'
# print(b in a)



#from bs4 import BeautifulSoup
#print(BeautifulSoup("&quot;Меня уже трясет. Кристиан наклоняется и целует меня между лопаток.<br>-Готова? <br>Готова? А готова ли я к такому?<br>-Да, - шепчу чуть слышно, едва ворочая сухим языком. Он сует что-то в меня. Черт, это же большой палец. Другие пальцы ласкают клитор. Я стону… от наслаждения. И пока одни пальцы творят это маленькое чудо, другие вводят в анус пробку. Кристиан замирает. Я слышу его хриплое, резкое дыхание и пытаюсь принять все ощущения: восхитительной полноты, тревожно-волнующей опасности, чисто эротическое наслаждение. Все они смешиваются, скручиваются в спирали, растекаются во мне. Кристиан осторожно нажимает на пробку…&quot; <br><br>***<br><br>Коллекция &quot;50 оттенков серого&quot;. Совершенно потрясающая анальная пробка из гладкого, нежнейшего силикона. Невероятная анальная стимуляция.<br><br>Идеальная анатомическая форма, зауженный кончик для легкого проникновения, т-образное основание, за которое удобно держаться, 3 скорости и семь вариантов вибрации. Ощущение наполненности и чувственное наслаждение.<br><br>Пробка водонепроницаема, ею можно играть даже в ванной. Вибрация: 1 батарейка ААА (в комплект не входит).<br><br>С пробкой можно использовать смазки на водной основе.", "lxml").text.replace("\xc2\xa0", " ").replace('&nbsp', '').replace('&quot', '').replace('…','').replace('*',''))


# a = '489080b144923'
# b = list(map(int,a))
# print(b)
# cats_wb = set()
# with open('./pool/SexOptovik/google_downloaded/wb_cats.txt') as f:
#     for lines in f:
#         cats_wb.add(lines)
#
# from itertools import product
# @staticmethod
# def init_category(data, cats_wb):
#     CONST_COUNT_WORDS = 3
#     data_lower = data.lower()
#     getting_in_cat_wb = set()
#     cats_wb = set()
#
#     info = list(map(lambda data: data.strip(), data_lower.lower().split('#')))
#     info = list(map(lambda info: info.split('>'), info))
#     with open('E:\csv_parser_wb\pool\SexOptovik\google_downloaded\wb_cats.txt') as f:
#         for i in f:
#             cats_wb.add(i.rstrip().lower())
#     cats_wb.remove('')
#
#     for i in info:
#         saved = []
#         for j in i:
#             if 'белье' not in j:
#                 for n in j.split():
#                     if n in cats_wb:
#                         getting_in_cat_wb.add(n)
#                 t = list(map(lambda j: j + ' ', j.split()))
#                 for p in saved:
#                     t.append(p + ' ')
#                 for n in range(CONST_COUNT_WORDS):
#                     t.append('')
#                 if 'эротик' not in t:
#                     t.append('эротик ')
#                 all_var = set(map(''.join, product(t, repeat=CONST_COUNT_WORDS)))
#                 all_var = list(map(lambda i: i.rstrip().replace(',', ''), all_var))
#                 #all_var = list(map(lambda i: i, all_var))
#                 for var in all_var:
#                     if var in cats_wb:
#                         getting_in_cat_wb.add(var)
#                 saved = j.split().copy()
#             else:
#                 t = data_lower.lower()
#                 ed = ['комплекты', 'бдсм', 'топы', 'лифы', 'бюстье', 'стрэпы', 'корсаж', ' маски',' мастурбаторы']
#                 ed_cats = ['комплекты эротик', 'комплекты бдсм', 'топы эротик', 'бюстгальтеры эротик',
#                            'бюстгальтеры эротик', 'стрэпы эротик', 'корсеты эротик','маски эротик','мастурбаторы мужские']
#                 for o in range(len(ed)):
#                     if ed[o] in t:
#                         getting_in_cat_wb.add(ed_cats[o])
#                 if 'трусики' in t or 'трусы' in t or 'стринги' in t or 'шортики' in t or 'шорты' in t:
#                     getting_in_cat_wb.add('трусы эротик')
#     if 'вибро' in data_lower:
#         ed = ['зажим', 'яйца', 'трус', 'пуля']
#         ed_cats = ['зажимы для сосков', 'виброяйца', 'вибротрусики', 'вибропули']
#         for o in range(len(ed)):
#             if ed[o] in data_lower:
#                 getting_in_cat_wb.add(ed_cats[o])
#
#     ed = ['лубрикан', 'стек ', 'смазки для секс-игрушек', 'наборы и аксессуары', 'клиторальные стимуляторы',
#           'костюмы', 'бдсм товары и фетиш > аксессуары', 'наручник', 'кольцо эрекционное',
#           'стимулирующие влагалище',
#           ' танго', 'стреп','стрэп', 'чулки',' пояс','мастурбатор']
#     ed_cats = ['лубриканты', 'стэки эротик', 'уходовые средства эротик', 'наборы игрушек для взрослых', 'вибраторы',
#                'ролевые костюмы эротик', 'комплекты бдсм', 'наручники эротик', 'эрекционные кольца',
#                'вагинальные тренажеры', 'трусы эротик','стрэпы эротик','стрэпы эротик','чулки эротик','пояса эротик','мастурбаторы мужские']
#     for o in range(len(ed)):
#         if ed[o] in data_lower:
#             getting_in_cat_wb.add(ed_cats[o])
#     res = []
#     if len(getting_in_cat_wb) > 1 and 'комплекты бдсм' in getting_in_cat_wb:
#         getting_in_cat_wb.remove('комплекты бдсм')
#     for i in getting_in_cat_wb:
#         cat = i[:1].upper() + i[1:]
#         if 'бдсм' in i:
#             cat = cat.replace('бдсм', 'БДСМ')
#         res.append(cat)
#     if len(res) == 0:
#         res.append('Наборы игрушек для взрослых')
#     return res[0], data.replace(' #', '.')
#
# b = 'Мастурбатор с вибрацией, подогревом и'
# print(init_category(b,cats_wb))

a = {'id-11668-1168-1', 'id-6535-1168'}
print('11668' in a)