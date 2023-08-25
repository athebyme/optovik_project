from __future__ import print_function

import os
import sys
from functools import lru_cache

# print(input())
import openpyxl
import pandas as pd
import pymorphy2
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.neighbors import NearestNeighbors

from src.AnyOtherCode import main
from src.dataConfigs.regEx import rePattern


# xlsx = openpyxl.load_workbook(os.getcwd()+'\\report_2022_7_29.xlsx')
# sheet = xlsx.active
# col_names= []
#
# for row in sheet.iter_rows(1,sheet.max_row):
#     print(row[3].value)
# set_of_post_codes = [
#     'Z1C1A0t-', 'W1C1Ayt', 'Z1C1L', 'Z1C1K', 'W1C1K', 'W1C1V', 'Z1C1A', 'W1C1L', 'W1C1A'
# ]
# item = '1234Z1C1L9645'
# if any(set_of_post_codes) not in item:
#     pass
# else:
#     print(set_of_post_codes.index(item))
# from pyfiglet import Figlet
#
# @staticmethod
# def showText(text):
#     try:
#         view_text = Figlet(font='slant')
#         print(view_text.renderText(text))
#     except:
#         print(text)
#
#
# showText('DNIS')
#
# def download_file(real_file_id):
#     """Downloads a file
#     Args:
#         real_file_id: ID of the file to download
#     Returns : IO object with location.
#
#     Load pre-authorized user credentials from the environment.
#     TODO(developer) - See https://developers.google.com/identity
#     for guides on implementing OAuth2 for the application.
#     """
#     creds, _ = google.auth.default()
#
#     try:
#         # create drive api client
#         service = build('drive', 'v3', credentials=creds)
#
#         file_id = real_file_id
#
#         # pylint: disable=maybe-no-member
#         request = service.files().get_media(fileId=file_id)
#         file = io.BytesIO()
#         downloader = MediaIoBaseDownload(file, request)
#         done = False
#         while done is False:
#             status, done = downloader.next_chunk()
#             print(F'Download {int(status.progress() * 100)}.')
#
#     except HttpError as error:
#         print(F'An error occurred: {error}')
#         file = None
#
#     return file.getvalue()
#
#
# if __name__ == '__main__':
#     download_file(real_file_id='1iOAtmIzmDJZZ1vRkWC0tOiyod6uP5ZugRI1hoJhqXjc')
# from itertools import product
#
# @staticmethod
# def init_category(data):
#     CONST_COUNT_WORDS = 3
#     data_lower = data.lower()
#     getting_in_cat_wb = set()
#     cats_wb = set()
#     info = list(map(lambda data: data.strip(), data_lower.lower().split('#')))
#     info = list(map(lambda info: info.split('>'), info))
#     with open('E:\csv_parser_wb\pool\SexOptovik\google_downloaded\wb_cats.txt') as f:
#         for i in f:
#             cats_wb.add(i.rstrip().lower())
#     cats_wb.remove('')
#
#     for i in info:
#         saved = []
#         for j in i:
#             if 'белье' not in j:
#                 for n in j.split():
#                     if n in cats_wb:
#                         getting_in_cat_wb.add(n)
#                 t = list(map(lambda j: j+' ', j.split()))
#                 for p in saved:
#                     t.append(p+' ')
#                 for n in range(CONST_COUNT_WORDS):
#                     t.append('')
#                 if 'эротик' not in t:
#                     t.append('эротик ')
#                 all_var = set(map(''.join,product(t,repeat = CONST_COUNT_WORDS)))
#                 all_var = list(map(lambda i:i.rstrip(),all_var))
#                 all_var = list(map(lambda i: i.replace(',',''),all_var))
#                 for var in all_var:
#                     if var in cats_wb:
#                         getting_in_cat_wb.add(var)
#                 saved = j.split().copy()
#             else:
#                 t = data_lower.lower()
#                 if 'комплекты' in t:
#                     getting_in_cat_wb.add('комплекты эротик')
#                 if 'бдсм' in t:
#                     getting_in_cat_wb.add('комплекты бдсм')
#                 if 'топы' in t:
#                     getting_in_cat_wb.add('топы эротик')
#                 if 'лифы' in t:
#                     getting_in_cat_wb.add('бюстгальтеры эротик')
#                 if 'бюстье' in t:
#                     getting_in_cat_wb.add('бюстгальтеры эротик')
#                 if 'стрэпы' in t:
#                     getting_in_cat_wb.add('стрэпы эротик')
#                 if 'корсаж' in t:
#                     getting_in_cat_wb.add('корсеты эротик')
#                 if 'трусики' in t or 'трусы' in t or 'стринги' in t or 'шортики' in t or 'шорты' in t:
#                     getting_in_cat_wb.add('трусы эротик')
#                 if 'маски' in t:
#                     getting_in_cat_wb.add('маски эротик')
#     if 'вибро' in data_lower:
#         if 'зажим' in data_lower:
#             getting_in_cat_wb.add('зажимы для сосков')
#         if 'яйца' in data_lower:
#             getting_in_cat_wb.add('виброяйца')
#         if 'трус' in data_lower:
#             getting_in_cat_wb.add('вибротрусики')
#         if 'пуля' or 'пули' in data_lower:
#             getting_in_cat_wb.add('вибропули')
#     ed = ['лубрикан', 'стек ', 'смазки для секс-игрушек', 'наборы и аксессуары', 'клиторальные стимуляторы',
#            'костюмы', 'бдсм товары и фетиш > аксессуары', 'наручник', 'кольцо эрекционное', 'стимулирующие влагалище',
#            ' танго', 'стреп']
#     cats_ed = ['лубриканты', 'стэки эротик', 'уходовые средства эротик', 'наборы игрушек для взрослых', 'вибраторы',
#                'ролевые костюмы эротик', 'комплекты бдсм', 'наручник эротик', 'эрекционные кольца',
#                'вагинальные тренажеры', 'трусы эротик','стрэпы эротик']
#     for o in range(len(ed)):
#         if ed[o] in data_lower:
#             getting_in_cat_wb.add(cats_ed[o])
#     res = []
#     if len(getting_in_cat_wb)>1 and 'комплекты бдсм' in getting_in_cat_wb:
#         getting_in_cat_wb.remove('комплекты бдсм')
#     for i in getting_in_cat_wb:
#         cat = i[:1].upper() + i[1:]
#         if 'БДСМ' in i:
#             cat = cat.replace('бдсм','БДСМ')
#         res.append(cat)
#     if len(res) == 0:
#         res.append('Наборы игрушек для взрослых')
#     return res[0], data.replace(' #','.')
#
# print(init_category('Красный блесиящий стреп H.E.L. Essie'))
# print(os.path.exists('./pool/SexOptovik/google_downloaded/wb_cats2.txt'))
#
#
# country = 'Китай-США' # Страна производства
# if '-' in country:
#     country = country[:country.find('-')]
# print(country)
# import re
#
# @staticmethod
# def for_sizes_parse(array_4XL):
#     size = ['XS','S','M','L','XL','XXL','XXXL','4XL']
#     size_ru = array_4XL.copy()
#     dict = {}
#     for i in range(len(size)):
#         dict.setdefault(size[i],size_ru[i])
#     return dict
# def parse_sizes(text,type):
#     data = list(map(str,text.split(', ')))
#     res = {}
#     if type == 'all':
#         for item in data:
#             it = item
#             if ' мм' in it:
#                 k = 10
#             elif ' кг' in it:
#                 k = 0.001
#             else:
#                 k = 1
#
#             #if '-' in it:
#             #    it = it[it.find('-')+1:]
#             if 'вес' in it and res.setdefault('weight') is None:
#                 w = 0
#                 if text.lower().count('вес') > 1:
#                     for i in data:
#                         if 'вес' in i:
#                             w += round(float(re.sub("[^0-9,]", "", i))/k,2)
#                 else:
#                     w = round(int(re.sub("[^0-9,]", "", it))/k,2)
#                 res['weight'] = w
#
#             eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
#             eds_en = ['length', 'width', 'height', 'diameter', 'depth']
#             for i in range(len(eds)):
#                 if eds[i] in it and res.setdefault(eds_en[i]) is None:
#                     _r = re.sub("[^0-9,-]", "", it).replace(',','.')
#                     if '-' in r:
#                         r_lst = list(map(str,r.split('-')))
#                         av = 0
#                         for j in r_lst:
#                             av += float(j)
#                         _r = round(float(av/len(r_lst)) / k, 2)
#                     else:
#                         _r = round(float(_r) / k, 2)
#                     res[eds_en[i]] = _r
#                     break
#         return res
#     if type == 'clothe':
#         SIZES = {1:'XS',2:'S',3:'M',4:'L',5:'XL',6:'XXL',7:'XXXL',8:'4XL'}
#         SIZES_RU = {'XS':'40-42','S':'42-44','M':'44-46','L':'46-48','XL':'48-50','1XL':'52-54','XXL':'54-56',
#                     'XXXL':'56-58','4XL':'58-60','XL':'50-52'}
#         for k,v in SIZES_RU.items():
#             if v in text:
#                 res['size'] = k
#                 res['size-ru'] = v
#         check = list(map(str,text.split(', ')))
#         for i in check:
#             if 'об. ' in i and res.setdefault('size') is None:
#                 n = re.sub("[^0-9-]", "", i)
#                 l = list(map(int, n.split('-')))
#                 sr_arf = round((l[1] + l[0]) / 2)
#                 if 'груди' in i:
#                     SIZE = for_sizes_parse([99,103,107,111,115,119,123,127])
#                 if 'бедер' in i:
#                     SIZE = for_sizes_parse([103,107,111,115,119,123,127,131])
#                 if 'талия' in i:
#                     SIZE = for_sizes_parse([86,90,94,98,102,107,112,117])
#                 if sr_arf < SIZE['XS']:
#                     res['size'] = 'XS'
#                     res['size-ru'] = SIZES_RU['S']
#                     return res
#                 prev_sz = SIZE['XS']
#                 for k,v in SIZE.items():
#                     if k == 'XS': pass
#                     if sr_arf < v and sr_arf > prev_sz:
#                         res['size'] = k
#                         res['size-ru'] = SIZES_RU[k]
#                         return res
#                     prev_sz = v
#
#         t = check[0]
#         if t.count('.') > 1:
#             res['size'] = 'M'
#             res['size-ru'] = SIZES_RU['M']
#             return res
#         if 'универсальный' in t:
#             res['size'] = 'M'
#             res['size-ru'] = SIZES_RU['M']
#             return res
#         if '-' in t and len(t) == 5:
#             res['size-ru'] = t
#             size = t.split('-')[0]
#             for k,v in SIZES_RU.items():
#                 if size in v:
#                     res['size'] = k
#             return res
#         for k,v in SIZES.items():
#             if (v in t or str(k) in t) and res.setdefault('size') is None:
#                 res['size'] = v
#                 res['size-ru'] = SIZES_RU[v]
#                 return res
#         return res
#     if type == 'cosmetic':
#         ed = ''
#         if ' мл' in text:
#             ed = 'мл'
#         if ' г' in text:
#             ed = 'г'
#         volume = re.sub("[^0-9 *хx,]", "", text)
#         volume = volume.replace('  ', ' *', 1)
#         res.setdefault('volume', volume.strip())
#         res.setdefault('ed', ed.strip())
#         return res
#
# size = parse_sizes('общая длина 23,5 см, диаметр 5,7-9 см',type='all')
# r = []
# r.append(size.get('123'))
# print(size.get('depth'), r)
# СОХРАНИТЬ
# @staticmethod
# def init_category(data, cats_wb):
#     CONST_COUNT_WORDS = 3
#     data_lower = data.lower()
#     getting_in_cat_wb = set()
#     cats_wb = set()
#     info = list(map(lambda data: data.strip(), data_lower.lower().split('#')))
#     info = list(map(lambda info: info.split('>'), info))
#     with open('E:\csv_parser_wb\pool\SexOptovik\google_downloaded\wb_cats.txt') as f:
#         for i in f:
#             cats_wb.add(i.rstrip().lower())
#     cats_wb.remove('')
#     if 'вибро' in data_lower:
#         if 'зажим' in data_lower:
#             getting_in_cat_wb.add('зажимы для сосков')
#         if 'яйца' in data_lower:
#             getting_in_cat_wb.add('виброяйца')
#         if 'трус' in data_lower:
#             getting_in_cat_wb.add('вибротрусики')
#         if 'пуля' or 'пули' in data_lower:
#             getting_in_cat_wb.add('вибропули')
#     if 'смазки для секс-игрушек' in data_lower:
#         getting_in_cat_wb.add('уходовые средства эротик')
#     if 'наборы и аксессуары' in data_lower:
#         getting_in_cat_wb.add('наборы игрушек для взрослых')
#     if 'клиторальные стимуляторы' in data_lower:
#         getting_in_cat_wb.add('вибраторы')
#     if 'костюмы' in data_lower:
#         getting_in_cat_wb.add('ролевые костюмы эротик')
#     if 'бдсм товары и фетиш > аксессуары' in data_lower:
#         getting_in_cat_wb.add('комплекты бдсм')
#     if 'наручник' in data_lower:
#         getting_in_cat_wb.add('наручник эротик')
#     if 'презерватив' in data_lower:
#         getting_in_cat_wb.add('презервативы')
#     for i in info:
#         saved = []
#         for j in i:
#             if 'белье' not in j:
#                 for n in j.split():
#                     if n in cats_wb:
#                         getting_in_cat_wb.add(n)
#                 t = list(map(lambda j: j + ' ', j.split()))
#                 for p in saved:
#                     t.append(p + ' ')
#                 for n in range(CONST_COUNT_WORDS):
#                     t.append('')
#                 # t.append('эротик')
#                 all_var = list(map(''.join, product(t, repeat=CONST_COUNT_WORDS)))
#                 all_var = list(map(lambda i: i.rstrip(), all_var))
#                 all_var = list(map(lambda i: i.replace(',', ''), all_var))
#                 for var in all_var:
#                     if var in cats_wb:
#                         getting_in_cat_wb.add(var)
#                 saved = j.split().copy()
#             else:
#                 t = data_lower.lower()
#                 if 'комплекты' in t:
#                     getting_in_cat_wb.add('комплекты эротик')
#                 if 'бдсм' in t:
#                     getting_in_cat_wb.add('комплекты бдсм')
#                 if 'топы' in t:
#                     getting_in_cat_wb.add('топы эротик')
#                 if 'лифы' in t:
#                     getting_in_cat_wb.add('бюстгальтеры эротик')
#                 if 'бюстье' in t:
#                     getting_in_cat_wb.add('бюстгальтеры эротик')
#                 if 'стрэпы' in t:
#                     getting_in_cat_wb.add('стрэпы эротик')
#                 if 'корсаж' in t:
#                     getting_in_cat_wb.add('корсеты эротик')
#                 if 'трусики' in t or 'трусы' in t or 'стринги' in t or 'шортики' in t or 'шорты' in t:
#                     getting_in_cat_wb.add('трусы эротик')
#                 if 'маски' in t:
#                     getting_in_cat_wb.add('маски эротик')
#     res = []
#     if len(getting_in_cat_wb) > 1 and 'комплекты бдсм' in getting_in_cat_wb:
#         getting_in_cat_wb.remove('комплекты бдсм')
#     for i in getting_in_cat_wb:
#         cat = i[:1].upper() + i[1:]
#         if 'БДСМ' in i:
#             cat = cat.replace('бдсм', 'БДСМ')
#         res.append(cat)
#     if len(res) == 0:
#         res.append('Наборы игрушек для взрослых')
#     return res[0], data.replace(' #', '.')
# def start(text):
#     if ', ' not in text:
#         _split_parametr = '  '
#         eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
#         indx = []
#         data = []
#         for x in eds:
#             if text.find(x) != -1: indx.append(text.find(x))
#         indx.sort()
#         if len(indx) > 1:
#             for x in range(1,len(indx)):
#                 data.append(text[:indx[x]-1].strip())
#                 text = text[indx[x]:]
#             data.append(text.strip())
#         else:
#             data.append(text)
#         return data
# print(start('длина 17,75 см,диаметр 11,5 см'))
# import re
# @staticmethod
# def for_sizes_parse(array_4XL):
#     size = ['XS', 'S', 'M', 'L', 'XL', 'XXL', 'XXXL', '4XL']
#     size_ru = array_4XL.copy()
#     dict = {}
#     for i in range(len(size)):
#         dict.setdefault(size[i], size_ru[i])
#     return dict
# @staticmethod
# def parse_sizes(text, type):
#     t = text
#     data = []
#     eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина', 'вес']
#     indx = []
#     indx_first = []
#     data = []
#     for x in eds:
#         if t.find(x) != -1: indx.append(t.find(x))
#     indx_first = indx.copy()
#     if len(indx) > 0:
#         if indx[0] != 0:
#             t = t[indx[0]:]
#             indx = []
#             for x in eds:
#                 if t.find(x) != -1: indx.append(t.find(x))
#         indx.sort()
#     if len(indx) != 0:
#         if len(indx) > 1:
#             for x in range(1, len(indx)):
#                 data.append(t[:indx[x] - 1].strip())
#                 t = t[indx[x]:]
#             data.append(t.strip())
#         else:
#             if text.strip() not in data:
#                 data.append(text.strip())
#     res = {}
#     if type == 'all':
#         for item in data:
#             it = item
#             if ' мм' in it:
#                 k = 10
#             elif ' кг' in it:
#                 k = 0.001
#             else:
#                 k = 1
#             # it = it.replace(',', '.')
#             # if '-' in it:
#             #    it = it[it.find('-')+1:]
#             if 'вес' in it and res.setdefault('weight') is None:
#                 it = it.replace(',', '.')
#                 w = 0
#                 if text.lower().count('вес') > 1:
#                     for i in data:
#                         if 'вес' in i:
#                             w += round(float(re.sub("[^0-9,]", "", i).replace(',','.')) / k, 2)
#                 else:
#                     w = round(float(re.sub("[^0-9,]", "", it).replace(',','.')) / k, 2)
#                 res['weight'] = w
#
#             eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
#             eds_en = ['length', 'width', 'height', 'diameter', 'depth']
#             for i in range(len(eds)):
#                 if eds[i] in it and res.setdefault(eds_en[i]) is None:
#                     r = re.sub("[^0-9,-]", "", it).replace(',', '.')
#                     if r == '':
#                         res[eds_en[i]] = 'универсальный (растягивается)'
#                         break
#                     if '-' in r:
#                         r_lst = list(map(str, r.split('-')))
#                         try:
#                             r_lst.remove('')
#                         except ValueError:
#                             pass
#                         av = 0
#                         for j in r_lst:
#                             av += float(j)
#                         r = round(float(av / len(r_lst)) / k, 2)
#                     else:
#                         r = round(float(r) / k, 2)
#                     res[eds_en[i]] = r
#                     break
#         return res
#     if type == 'clothe':
#         SIZES = {1: 'XS', 2: 'S', 3: 'M', 4: 'L', 5: 'XL', 6: 'XXL', 7: 'XXXL', 8: '4XL'}
#         SIZES_RU = {'XS': '40-42', 'S': '42-44', 'M': '44-46', 'L': '46-48', 'XL': '48-50', '1XL': '52-54',
#                     'XXL': '54-56',
#                     'XXXL': '56-58', '4XL': '58-60', 'XL': '50-52'}
#         eds = ['об. талии', 'об. груди', 'об. бедер']
#         indx_obhvats = []
#         for i in eds:
#             if i in text:
#                 indx_obhvats.append(text.index(i))
#         if len(data) != 0:
#             eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
#             eds_en = ['length', 'width', 'height', 'diameter', 'depth']
#             for x in data:
#                 temp = x
#                 if ' мм' in temp:
#                     k = 10
#                 elif ' кг' in temp:
#                     k = 0.001
#                 else:
#                     k = 1
#                 for z in range(len(eds)):
#                     if eds[z] in x:
#                         t = re.sub("[^0-9,-]", "", x).replace(',', '.')
#                         if res.setdefault(eds_en[z]) is None:
#                             res[eds_en[z]] = t
#                 if 'вес' in temp and res.setdefault('weight') is None:
#                     it = temp.replace(',', '.')
#                     w = 0
#                     if text.lower().count('вес') > 1:
#                         for i in data:
#                             if 'вес' in i:
#                                 w += round(float(re.sub("[^0-9,]", "", i)) / k, 2)
#                     else:
#                         w = round(float(re.sub("[^0-9,]", "", it)) / k, 2)
#                     res['weight'] = w
#
#         for k, v in SIZES_RU.items():
#             if v in text:
#                 res['size'] = k
#                 res['size-ru'] = v
#         if len(indx_obhvats) != 0 and len(indx_first) != 0 :
#             if indx_first[len(indx_first) - 1] > indx_obhvats[len(indx_obhvats) - 1]:
#                 text = text[:indx_first[0]]
#             else:
#                 pass
#         check = list(map(str, text.split(', ')))
#
#         for i in check:
#             if 'об. ' in i and res.setdefault('size') is None:
#                 n = re.sub("[^0-9-]", "", i)
#                 l = list(map(int, n.split('-')))
#                 sr_arf = round((l[1] + l[0]) / 2)
#                 if 'груд' in i:
#                     SIZE = for_sizes_parse([99, 103, 107, 111, 115, 119, 123, 127])
#                 if 'бедер' in i:
#                     SIZE = for_sizes_parse([103, 107, 111, 115, 119, 123, 127, 131])
#                 if 'тали' in i:
#                     SIZE = for_sizes_parse([86, 90, 94, 98, 102, 107, 112, 117])
#                 if sr_arf < SIZE['XS']:
#                     res['size'] = 'XS'
#                     res['size-ru'] = SIZES_RU['S']
#                     return res
#                 prev_sz = SIZE['XS']
#                 for k, v in SIZE.items():
#                     if k == 'XS': pass
#                     if sr_arf < v and sr_arf > prev_sz:
#                         res['size'] = k
#                         res['size-ru'] = SIZES_RU[k]
#                         return res
#                     prev_sz = v
#
#         t = check[0]
#         if t.count('.') > 1:
#             res['size'] = 'M'
#             res['size-ru'] = SIZES_RU['M']
#             return res
#         if 'универсальный' in t:
#             res['size'] = 'M'
#             res['size-ru'] = SIZES_RU['M']
#             return res
#         if '-' in t and len(t) == 5:
#             res['size-ru'] = t
#             size = t.split('-')[0]
#             for k, v in SIZES_RU.items():
#                 if size in v:
#                     res['size'] = k
#             return res
#         for k, v in SIZES.items():
#             if (v in t or str(k) in t) and res.setdefault('size') is None:
#                 res['size'] = v
#                 res['size-ru'] = SIZES_RU[v]
#                 return res
#         return res
#     if type == 'cosmetic':
#         ed = ''
#         if ' мл' in text:
#             ed = 'мл'
#         if ' г' in text:
#             ed = 'г'
#         volume = re.sub("[^0-9 *хx,]", "", text)
#         volume = volume.replace('  ', ' *', 1)
#         res.setdefault('volume', volume.strip())
#         res.setdefault('ed', ed.strip())
#         return res
#     return text
# print(parse_sizes('диаметр шариков  - 2,8 см, вес белого - 14 г, вес серого - 29 г, вес темно-серого 36 г, вес черного 55 г',type = 'all'))
# a = ''
# a = float(10)
# print(a)


# path = './shrih/Список штрихкодов (4).xlsx'
# #with ope n
# with open(path,encoding='ansi') as f:
#     for lines in f:
#         print(lines)


# @staticmethod
# def parse_sizes(text, type):
#     eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
#     eds_en = ['length', 'width', 'height', 'diameter', 'depth']
#     if ', ' not in text or type == 'clothe':
#         t = text
#         data = []
#         eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина', 'вес']
#         indx = []
#         indx_first = []
#         data = []
#         for x in eds:
#             if t.find(x) != -1: indx.append(t.find(x))
#         indx.sort()
#         indx_first = indx.copy()
#         if len(indx) != 0:
#             if indx[0] != 0:
#                 t = t[indx[0]:]
#                 indx = []
#                 for x in eds:
#                     if t.find(x) != -1: indx.append(t.find(x))
#
#             if len(indx) > 1:
#                 for x in range(1, len(indx)):
#                     data.append(t[:indx[x] - 1].strip())
#                     t = t[indx[x]:]
#                 data.append(t.strip())
#             else:
#                 if text.strip() not in data:
#                     data.append(text.strip())
#     else:
#         _split_parametr = ', '
#         data = list(map(str, text.split(_split_parametr)))
#     res = {}
#     ed_local_check = [',длина', ',ширина', ',высота', ',диаметр', ',глубина']
#     for i in data:
#         FLAG = []
#         for b in ed_local_check:
#             if b in i:
#                 FLAG.append(True)
#         if len(FLAG) > 1:
#             _indx = []
#             divided = []
#             for j in range(len(ed_local_check)):
#                 if i.find(eds[j]) != -1: _indx.append(i.find(eds[j]))
#             _indx.sort()
#             data.remove(i)
#             t = i
#             for x in range(1, len(_indx)):
#                 data.append(t[:_indx[x] - 1].strip())
#                 t = t[_indx[x]:]
#             data.append(t.strip())
#
#     ed_local_check = [',длина', ',ширина', ',высота', ',диаметр', ',глубина']
#     for i in data:
#         FLAG = []
#         for b in ed_local_check:
#             if b in i:
#                 FLAG.append(True)
#         if len(FLAG) >= 1:
#             _indx = []
#             divided = []
#             for j in range(len(ed_local_check)):
#                 if i.find(eds[j]) != -1: _indx.append(i.find(eds[j]))
#             _indx.sort()
#             data.remove(i)
#             t = i
#             for x in range(1, len(_indx)):
#                 data.append(t[:_indx[x] - 1].strip())
#                 t = t[_indx[x]:]
#             data.append(t.strip())
#     if type == 'all':
#         for item in data:
#             it = item
#             if ' мм' in it:
#                 k = 10
#             elif ' кг' in it:
#                 k = 0.001
#             else:
#                 k = 1
#             # it = it.replace(',', '.')
#             # if '-' in it:
#             #    it = it[it.find('-')+1:]
#             if 'вес' in it and res.setdefault('weight') is None:
#                 it = it.replace(',', '.')
#                 w = 0
#                 if text.lower().count('вес') > 1:
#                     for i in data:
#                         if 'вес' in i:
#                             w += round(float(re.sub("[^0-9,]", "", i).replace(',', '.')) / k, 2)
#                 else:
#                     w = round(float(re.sub("[^0-9,]", "", it).replace(',', '.')) / k, 2)
#                 res['weight'] = w
#
#             for i in range(len(eds)):
#                 if eds[i] in it and res.setdefault(eds_en[i]) is None:
#                     r = re.sub("[^0-9,-]", "", it).replace(',', '.')
#                     if r == '':
#                         res[eds_en[i]] = 'универсальный (растягивается)'
#                         break
#                     if '-' in r:
#                         r_lst = list(map(str, r.split('-')))
#                         try:
#                             r_lst.remove('')
#                         except ValueError:
#                             pass
#                         av = 0
#                         for j in r_lst:
#                             av += float(j)
#                         r = round(float(av / len(r_lst)) / k, 2)
#                     else:
#                         r = round(float(r) / k, 2)
#                     res[eds_en[i]] = r
#                     break
#         return res
#     if type == 'clothe':
#         SIZES = {1: 'XS', 2: 'S', 3: 'M', 4: 'L', 5: 'XL', 6: 'XXL', 7: 'XXXL', 8: '4XL'}
#         SIZES_RU = {'XS': '40-42', 'S': '42-44', 'M': '44-46', 'L': '46-48', 'XL': '48-50', '1XL': '52-54',
#                     'XXL': '54-56',
#                     'XXXL': '56-58', '4XL': '58-60', 'XL': '50-52'}
#         eds = ['об. талии', 'об. груди', 'об. бедер']
#         indx_obhvats = []
#         for i in eds:
#             if i in text:
#                 indx_obhvats.append(text.index(i))
#         if len(data) != 0:
#             eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
#             eds_en = ['length', 'width', 'height', 'diameter', 'depth']
#             for x in data:
#                 temp = x
#                 if ' мм' in temp:
#                     k = 10
#                 elif ' кг' in temp:
#                     k = 0.001
#                 else:
#                     k = 1
#                 for z in range(len(eds)):
#                     if eds[z] in x:
#                         t = re.sub("[^0-9,-]", "", x).replace(',', '.')
#                         if res.setdefault(eds_en[z]) is None:
#                             res[eds_en[z]] = t
#                 if 'вес' in temp and res.setdefault('weight') is None:
#                     it = temp.replace(',', '.')
#                     w = 0
#                     if text.lower().count('вес') > 1:
#                         for i in data:
#                             if 'вес' in i:
#                                 w += round(float(re.sub("[^0-9,]", "", i)) / k, 2)
#                     else:
#                         w = round(float(re.sub("[^0-9,]", "", it)) / k, 2)
#                     res['weight'] = w
#
#         for k, v in SIZES_RU.items():
#             if v in text:
#                 res['size'] = k
#                 res['size-ru'] = v
#         if len(indx_obhvats) != 0 and len(indx_first) != 0:
#             if indx_first[len(indx_first) - 1] > indx_obhvats[len(indx_obhvats) - 1]:
#                 text = text[:indx_first[0]]
#             else:
#                 pass
#         check = list(map(str, text.split(', ')))
#
#         for i in check:
#             if 'об. ' in i and res.setdefault('size') is None:
#                 n = re.sub("[^0-9-]", "", i)
#                 l = list(map(int, n.split('-')))
#                 sr_arf = round((l[1] + l[0]) / 2)
#                 if 'груд' in i:
#                     SIZE = for_sizes_parse([99, 103, 107, 111, 115, 119, 123, 127])
#                 if 'бедер' in i:
#                     SIZE = for_sizes_parse([103, 107, 111, 115, 119, 123, 127, 131])
#                 if 'тали' in i:
#                     SIZE = for_sizes_parse([86, 90, 94, 98, 102, 107, 112, 117])
#                 if sr_arf < SIZE['XS']:
#                     res['size'] = 'XS'
#                     res['size-ru'] = SIZES_RU['S']
#                     return res
#                 prev_sz = SIZE['XS']
#                 for k, v in SIZE.items():
#                     if k == 'XS': pass
#                     if sr_arf < v and sr_arf > prev_sz:
#                         res['size'] = k
#                         res['size-ru'] = SIZES_RU[k]
#                         return res
#                     prev_sz = v
#
#         t = check[0]
#         if t.count('.') > 1:
#             res['size'] = 'M'
#             res['size-ru'] = SIZES_RU['M']
#             return res
#         if 'универсальный' in t:
#             res['size'] = 'M'
#             res['size-ru'] = SIZES_RU['M']
#             return res
#         if '-' in t and len(t) == 5:
#             res['size-ru'] = t
#             size = t.split('-')[0]
#             for k, v in SIZES_RU.items():
#                 if size in v:
#                     res['size'] = k
#             return res
#         for k, v in SIZES.items():
#             if (v in t or str(k) in t) and res.setdefault('size') is None:
#                 res['size'] = v
#                 res['size-ru'] = SIZES_RU[v]
#                 return res
#         return res
#     if type == 'cosmetic':
#         ed = ''
#         if ' мл' in text:
#             ed = 'мл'
#         if ' г' in text:
#             ed = 'г'
#         volume = re.sub("[^0-9 *хx,]", "", text)
#         volume = volume.replace('  ', ' *', 1)
#         res.setdefault('volume', volume.strip())
#         res.setdefault('ed', ed.strip())
#         return res
#     return text
# print(parse_sizes(
#     'общая длина бутылки 28 см, макс. диаметр 7,2 см,глубина проникновения 14 см, внутренний диаметр 1,5 см (растягивается)',
#     type='all'))

# from itertools import product
#
# #t =
# CONST_COUNT_WORDS = 3
# all_var = set(map(''.join, product(t, repeat=CONST_COUNT_WORDS)))
# all_var = list(map(lambda i: i.rstrip().replace(',', ''), all_var))
# a = 'наручники эротик'
# b = 'наруч эротик'
# print(b in a)


# from bs4 import BeautifulSoup
# print(BeautifulSoup("&quot;Меня уже трясет. Кристиан наклоняется и целует меня между лопаток.<br>-Готова? <br>Готова? А готова ли я к такому?<br>-Да, - шепчу чуть слышно, едва ворочая сухим языком. Он сует что-то в меня. Черт, это же большой палец. Другие пальцы ласкают клитор. Я стону… от наслаждения. И пока одни пальцы творят это маленькое чудо, другие вводят в анус пробку. Кристиан замирает. Я слышу его хриплое, резкое дыхание и пытаюсь принять все ощущения: восхитительной полноты, тревожно-волнующей опасности, чисто эротическое наслаждение. Все они смешиваются, скручиваются в спирали, растекаются во мне. Кристиан осторожно нажимает на пробку…&quot; <br><br>***<br><br>Коллекция &quot;50 оттенков серого&quot;. Совершенно потрясающая анальная пробка из гладкого, нежнейшего силикона. Невероятная анальная стимуляция.<br><br>Идеальная анатомическая форма, зауженный кончик для легкого проникновения, т-образное основание, за которое удобно держаться, 3 скорости и семь вариантов вибрации. Ощущение наполненности и чувственное наслаждение.<br><br>Пробка водонепроницаема, ею можно играть даже в ванной. Вибрация: 1 батарейка ААА (в комплект не входит).<br><br>С пробкой можно использовать смазки на водной основе.", "lxml").text.replace("\xc2\xa0", " ").replace('&nbsp', '').replace('&quot', '').replace('…','').replace('*',''))


# a = '489080b144923'
# b = list(map(int,a))
# print(b)
# cats_wb = set()
# with open('./pool/SexOptovik/google_downloaded/wb_cats.txt') as f:
#     for lines in f:
#         cats_wb.add(lines)
#
# from itertools import product
# @staticmethod
# def init_category(data, cats_wb):
#     CONST_COUNT_WORDS = 3
#     data_lower = data.lower()
#     getting_in_cat_wb = set()
#     cats_wb = set()
#
#     info = list(map(lambda data: data.strip(), data_lower.lower().split('#')))
#     info = list(map(lambda info: info.split('>'), info))
#     with open('E:\csv_parser_wb\pool\SexOptovik\google_downloaded\wb_cats.txt') as f:
#         for i in f:
#             cats_wb.add(i.rstrip().lower())
#     cats_wb.remove('')
#
#     for i in info:
#         saved = []
#         for j in i:
#             if 'белье' not in j:
#                 for n in j.split():
#                     if n in cats_wb:
#                         getting_in_cat_wb.add(n)
#                 t = list(map(lambda j: j + ' ', j.split()))
#                 for p in saved:
#                     t.append(p + ' ')
#                 for n in range(CONST_COUNT_WORDS):
#                     t.append('')
#                 if 'эротик' not in t:
#                     t.append('эротик ')
#                 all_var = set(map(''.join, product(t, repeat=CONST_COUNT_WORDS)))
#                 all_var = list(map(lambda i: i.rstrip().replace(',', ''), all_var))
#                 #all_var = list(map(lambda i: i, all_var))
#                 for var in all_var:
#                     if var in cats_wb:
#                         getting_in_cat_wb.add(var)
#                 saved = j.split().copy()
#             else:
#                 t = data_lower.lower()
#                 ed = ['комплекты', 'бдсм', 'топы', 'лифы', 'бюстье', 'стрэпы', 'корсаж', ' маски',' мастурбаторы']
#                 ed_cats = ['комплекты эротик', 'комплекты бдсм', 'топы эротик', 'бюстгальтеры эротик',
#                            'бюстгальтеры эротик', 'стрэпы эротик', 'корсеты эротик','маски эротик','мастурбаторы мужские']
#                 for o in range(len(ed)):
#                     if ed[o] in t:
#                         getting_in_cat_wb.add(ed_cats[o])
#                 if 'трусики' in t or 'трусы' in t or 'стринги' in t or 'шортики' in t or 'шорты' in t:
#                     getting_in_cat_wb.add('трусы эротик')
#     if 'вибро' in data_lower:
#         ed = ['зажим', 'яйца', 'трус', 'пуля']
#         ed_cats = ['зажимы для сосков', 'виброяйца', 'вибротрусики', 'вибропули']
#         for o in range(len(ed)):
#             if ed[o] in data_lower:
#                 getting_in_cat_wb.add(ed_cats[o])
#
#     ed = ['лубрикан', 'стек ', 'смазки для секс-игрушек', 'наборы и аксессуары', 'клиторальные стимуляторы',
#           'костюмы', 'бдсм товары и фетиш > аксессуары', 'наручник', 'кольцо эрекционное',
#           'стимулирующие влагалище',
#           ' танго', 'стреп','стрэп', 'чулки',' пояс','мастурбатор']
#     ed_cats = ['лубриканты', 'стэки эротик', 'уходовые средства эротик', 'наборы игрушек для взрослых', 'вибраторы',
#                'ролевые костюмы эротик', 'комплекты бдсм', 'наручники эротик', 'эрекционные кольца',
#                'вагинальные тренажеры', 'трусы эротик','стрэпы эротик','стрэпы эротик','чулки эротик','пояса эротик','мастурбаторы мужские']
#     for o in range(len(ed)):
#         if ed[o] in data_lower:
#             getting_in_cat_wb.add(ed_cats[o])
#     res = []
#     if len(getting_in_cat_wb) > 1 and 'комплекты бдсм' in getting_in_cat_wb:
#         getting_in_cat_wb.remove('комплекты бдсм')
#     for i in getting_in_cat_wb:
#         cat = i[:1].upper() + i[1:]
#         if 'бдсм' in i:
#             cat = cat.replace('бдсм', 'БДСМ')
#         res.append(cat)
#     if len(res) == 0:
#         res.append('Наборы игрушек для взрослых')
#     return res[0], data.replace(' #', '.')
#
# b = 'Мастурбатор с вибрацией, подогревом и'
# print(init_category(b,cats_wb))

# a = {'id-11668-1168-1', 'id-6535-1168'}
# print('11668' in a)
#
# b = 4.4
# r = round(b + b*0.1,3)
# print(r)

# @staticmethod
# def clean_for_data(articular, seller_code, shortArticular=True):
#     temp = ''.join(articular)
#     articular = ''.join(articular)
#     original_code = seller_code
#     new_articular = current_code = ''
#     if shortArticular:
#         set_of_post_codes = [
#             'UNCOMMENT PREV ARRAY IF CHECK Z1 NEEDED'
#         ]
#     else:
#         set_of_post_codes = [
#             'Z1C1A0t-', 'W1C1Ayt', 'Z1C1L', 'Z1C1K', 'W1C1K', 'W1C1V', 'Z1C1A', 'W1C1L', 'W1C1A', 'Z1C1V'
#         ]
#     check_any_code_in = []
#     for i in set_of_post_codes:
#         check_any_code_in.append(i in articular)
#     if 'id' in articular:
#         articular = articular[articular.find('-') + 1:]
#         articular = articular[:articular.find('-')]
#         if 4 + len(seller_code) + len(articular) != len(temp):
#             return False, temp
#         new_articular = f'id-{articular}-{seller_code}'
#     elif any(check_any_code_in):
#         for i in set_of_post_codes:
#             if i in articular:
#                 current_code = i
#                 break
#         ind = articular.find(current_code)
#         if len(articular[:ind]) != len(seller_code):
#             seller_code = articular[:ind]
#         articular = articular[ind:]
#
#         c = len(current_code)
#         articular = articular.replace(current_code, '')
#
#         if len(articular) + c + len(seller_code) != len(temp):
#             return False, temp
#         new_articular = f'{original_code}{set_of_post_codes[check_any_code_in.index(True)]}{articular}'
#     else:
#         return False, temp
#     if shortArticular:
#         return True, articular
#     else:
#         return True, new_articular
# def init_stocks():
#     set_stocks = set()
#     with open(r"C:\Users\Anton\Documents\stock.csv", 'r+', encoding='utf-8') as stock:
#         for line in stock:
#             articul = clean_for_data(shortArticular=True, seller_code='1366', articular=line.rstrip().replace(';',''))
#             if articul[0]:
#                 set_stocks.add(articul[1])
#     return set_stocks
# print(len(init_stocks()))
def check_v(s):
    try:
        t = (list(map(int, s.split())))
        t[0] = t[0]
        for i in t:
            a = int(i)
            if len(str(a)) > 2:
                return 0, ''
        photo_str = s
        return 1, photo_str
    except Exception as ex:
        return 0, ''


def re_init_photo():
    a = main.Functions()
    dict = {}
    saver = {1: ['Артикул товара'], 2: ['Медиафайлы']}
    path = main.Functions.getFolderFile(0, item='товары без фото')
    set_a = a.getDataXslx(path=path, sellerCode='1277', _row=6)[0]
    with open('../../SexOptovik/all_prod_info.csv', 'r+') as f:
        for lines in f:
            s = ''
            DATA = list(map(lambda line: line.replace('"', ''), lines.split(';')))
            if DATA[0] in set_a:
                v = [0, '']
                data_row = 13
                while not v[0]:
                    v = check_v(DATA[data_row])
                    data_row += 1
                photo_urls = list(map(lambda
                                          photo_str: f'http://sexoptovik.ru/_project/user_images/prods_res/{DATA[0]}/{DATA[0]}_{photo_str}_{650}.jpg',
                                      v[1].split()))
                for i in range(len(photo_urls) - 1):
                    s += f'{photo_urls[i]}; '
                s += photo_urls[len(photo_urls) - 1]
                saver[1].append(f'id-{DATA[0]}-1366')
                saver[2].append(s)
        a.save_data(saver, seller_code='1366', path=r"C:\Users\Anton\Desktop")


def getFib(n):
    if n < 2: return n;
    return getFib(n - 1) + getFib(n - 2)


def photo(path='', _row=0, checkBrand='', seller_code=''):
    articuls = set()
    errors = set()
    lieBrands = set()
    xlsx = openpyxl.load_workbook(path)
    sheet = xlsx.active
    for row in sheet.iter_rows(2, sheet.max_row):
        articular = main.Functions.cleanArticul(row[_row].value, seller_code=seller_code)
        if articular[0]:
            articuls.add(articular[1])
            if row[0].value is not None and row[0].value.lower() in checkBrand.lower():
                lieBrands.add(articular[1])
        else:
            errors.add(articular[1])
    return articuls, errors, lieBrands


# set_a = main.Functions.getData(main.Functions.getFolderFile(0), 1366,)
# re_init_photo()
# saver = {1: ['Артикул товара'], 2: ['Медиафайлы']}
# articuls, errors, lieBrands = photo(r"C:\Users\Anton\Downloads\report_2022_12_2.xlsx (2).xlsx", 3, "BANANZZA", "1366")
# with open('./SexOptovik/all_prod_info.csv', 'r+') as f:
#     for lines in f:
#         s = ''
#         DATA = list(map(lambda line: line.replace('"', ''), lines.split(';')))
#         if DATA[0] in articuls:
#             v = [0, '']
#             data_row = 13
#             while not v[0]:
#                 v = check_v(DATA[data_row])
#                 data_row += 1
#             photo_urls = list(map(lambda
#                                       photo_str: f'http://sexoptovik.ru/_project/user_images/prods_res/{DATA[0]}/{DATA[0]}_{photo_str}_{650}.jpg',
#                                   v[1].split()))
#             for i in range(len(photo_urls) - 1):
#                 s += f'{photo_urls[i]}; '
#             s += photo_urls[len(photo_urls) - 1]
#             saver[1].append(f'id-{DATA[0]}-1366')
#             saver[2].append(s)
#     main.Functions().save_data(saver, seller_code='1366', path=r"C:\Users\Anton\Desktop")


def getConfig(sellerCode: str):
    conf = None
    v = -1
    if not os.path.isdir('../../config'): os.mkdir('../../config')
    if not os.path.isdir('../../config/shopConfs'): os.mkdir('../../config/shopConfs')
    if sellerCode == '1168':
        success = False
        while not success:
            try:
                v = int(input('Выберите магазин:\n[0] - Amare\n[1] - Somnium Face\n:'))
                if 0 <= v <= 1:
                    success = True
                else:
                    print('[!] Вы ввели недопустимое значение')
            except ValueError:
                print('[!] Введите число - 0 или 1')
        if v == 1:
            from config.shopConfs.SomniumfaceConfig import SomniumfaceConfig
            conf = SomniumfaceConfig()
        else:
            from config.shopConfs.AmareConfig import AmareConfig
            conf = AmareConfig()
    elif sellerCode == '1269':
        from config.shopConfs.LascivaConfig import LascivaConfig
        conf = LascivaConfig()
    elif sellerCode == '1292':
        from config.shopConfs.WisteriaConfig import WisteriaConfig
        conf = WisteriaConfig()
    elif sellerCode == '1366':
        from config.shopConfs.BananzzaConfig import BananzzaConfig
        conf = BananzzaConfig()
    return conf


import SexOptovik_ozon


def getCategoryFilesPaths(path):
    paths = [f for f in os.listdir(path) if f.endswith(".xlsx")]
    if len(paths) == 0:
        print(
            '[!] Проверьте файлы в папке шаблонов Ozon. Данных не найдено !\nВы можете попробовать заново загрузить файлы')
        sys.exit(1)
    return paths


def checkCharset(path):
    import chardet
    with open(path, 'rb') as f:
        result = chardet.detect(f.read())
    return result['encoding']


def download_msc():
    from src.AnyOtherCode.main import Functions as func
    main = func(abs_path='../../', provider='', sellerCode='')
    name = ['as_price_and_num_rrc.csv', 'ke_price_and_num_rrc.csv', 'an_price_and_num_rrc.csv']
    url = ['http://www.sexoptovik.ru/mp/as_price_and_num_rrc.csv',
           'http://www.sexoptovik.ru/mp/ke_price_and_num_rrc.csv',
           'http://www.sexoptovik.ru/mp/an_price_and_num_rrc.csv']
    if not os.path.exists('../../MSC'): os.mkdir('../../MSC')
    for i in range(len(name)):
        path = f'./MSC/{name[i]}'
        try:
            os.remove(path)
            main.download_universal(url=url[i], path_def='../../MSC')
        except OSError:
            main.download_universal(url=url[i], path_def='../../MSC')


def getArticulBrand(path, encoding):
    import re
    Articul_Brand = {}
    Photos = {}
    https = r'https://astkol.com'
    df = pd.read_csv(path, encoding=encoding, sep=';')
    if 'as_price_and_num_rrc' in path:
        df = df.iloc[0:]
        for i in range(len(df)):
            art = str(df.iloc[i, 0])
            Articul_Brand[art] = str(df.iloc[i, 11])
            url = df.iat[i, 6]
            first = url[url.find('//') + 2:]
            first = first[first.find('/'):]
            Photos[art] = [f'{https}{first}']
            if not pd.isna(df.iat[i, 40]):
                other = df.iat[i, 40].split(',')
                for j in other:
                    Photos[art].append(f'{https}{j}')
            totalStringUrl = ''
            for j in range(len(Photos.get(art))):
                totalStringUrl += Photos.get(art)[j]
                if j != len(Photos.get(art)) - 1:
                    totalStringUrl += '; '
            Photos[art] = totalStringUrl
    elif 'ke_price_and_num_rrc' in path:
        with open(path, 'r+', encoding='utf-8') as f:
            for line in f.readlines():
                print(line)
        df = df.iloc[22:]
        Articul_Brand = dict(zip(df.iloc[:, 3], df.iloc[:, 18]))
    elif 'an_price_and_num_rrc' in path:
        df = df.iloc[0:]
        for i in range(len(df)):
            art = ''
            if '-' in df.iloc[i, 0]:
                art = re.findall('-(\d+)', df.iloc[i, 0])[0]
                Articul_Brand[art] = df.iloc[i, 38]
            else:
                art = df.iloc[i, 0]
                Articul_Brand[df.iloc[i, 0]] = df.iloc[i, 4]
            Photos[art] = [f'{df.iat[i, 13]}']
            for j in range(14, 16):
                if not pd.isna(df.iat[i, j]):
                    Photos[art].append(df.iat[i, j])
            totalStringUrl = ''
            for j in range(len(Photos.get(art))):
                totalStringUrl += Photos.get(art)[j]
                if j != len(Photos.get(art)) - 1:
                    totalStringUrl += '; '
            Photos[art] = totalStringUrl
    return [Articul_Brand, Photos]


def changeBrands(path, brand):
    # download_msc()
    theOne = None
    newed = 0
    newed_loc = 0
    NEWED_ALL = 0
    changed = []
    andrey = []
    andrey_changed = []
    _ke = _id = _as = _an = 0
    ke_dict = {'----': 1}
    ke_photo = {}
    as_dict, as_photo = getArticulBrand(path='../../MSC/as_price_and_num_rrc.csv', encoding='cp1251')
    # ke_dict = getArticulBrand(path='./MSC/ke_price_and_num_rrc.csv', encoding='UTF-8')
    an_dict, an_photo = getArticulBrand(path='../../MSC/an_price_and_num_rrc.csv', encoding='cp1251')
    files = getCategoryFilesPaths(path=path)
    photo = ['']
    getAllMSC = None
    theOne = pd.DataFrame(
        {'Номенклатура': [''], 'Предмет': [''], 'Артикул товара': [''], 'Наименование': [''], 'Бренд': [''],
         'Описание': [''], 'Наличие фото': [''], 'Наличие видео': [''], 'Размер': [''], 'Рос. размер': [''],
         'Баркод товара': [''], 'Материал изделия': [''], 'Упаковка': [''], 'Вес товара без упаковки (г), г': [''],
         'Длина предмета, см': [''], 'Ширина упаковки, см': [''], 'Высота упаковки, см': [''],
         'Длина упаковки, см': [''], 'Количество предметов в упаковке': [''], 'Особенности секс игрушки': [''],
         'Диаметр секс игрушки': [''], 'Количество шариков': [''], 'Комплектация': [''], 'Цвет': [''],
         'Страна производства': [''], 'Номер декларации соответствия': [''], 'Номер сертификата соответствия': [''],
         'Дата регистрации сертификата/декларации': [''], 'Дата окончания действия сертификата/декларации': ['']})
    getAllMSC = theOne.copy()
    for file in files:
        newed_loc = 0
        df = pd.read_excel(fr'{path}\{file}', sheet_name='Sheet1')
        length = df.iloc[:, 4].tolist()
        for i in range(len(length)):
            try:
                column = 0
                for j in range(len(df.loc[0])):
                    if any(['1C1K' in str(df.iat[column, j]), '1C1A' in str(df.iat[column, j]),
                            '1C1L' in str(df.iat[column, j]), 'id-' in str(df.iat[column, j])]):
                        column = j
                        break
                if '1C1K' in df.iat[i, column]:
                    _ke += 1
                if 'id-' in df.iat[i, column]:
                    _id += 1
                if '1C1L' in df.iat[i, column]:
                    _as += 1
                if '1C1A' in df.iat[i, column]:
                    _an += 1
                    andrey.append(df.iat[i, column])
                articul = df.iat[i, column]
                articul = articul[articul.find('1C1') + 3:]
                if 't' in articul or '-' in articul:
                    while 't' in articul or '-' in articul:
                        articul = articul[1:]
                while not (ord(articul[0]) >= 48 and ord(articul[0]) <= 57):
                    articul = articul[1:]

                if any(['1C1A' in df.iat[i, column], '1C1L' in df.iat[i, column], '1C1K' in df.iat[i, column]]):
                    getAllMSC = getAllMSC.append(df.loc[i], ignore_index=True)

                if as_dict.get(articul) is not None and 'Z1C1L' in df.iat[i, column]:
                    df.iloc[i, 4] = as_dict.get(articul)
                    theOne = theOne.append(df.loc[i], ignore_index=True)
                    photo.append(as_photo.get(articul))
                    newed += 1
                    newed_loc += 1
                elif ke_dict.get(articul) is not None and 'Z1C1K' in df.iat[i, column]:
                    df.iloc[i, 4] = ke_dict.get(articul)
                    theOne = theOne.append(df.loc[i], ignore_index=True)
                    photo.append(ke_photo.get(articul))
                    newed += 1
                    newed_loc += 1
                elif an_dict.get(articul) is not None and 'Z1C1A' in df.iat[i, column]:
                    df.iloc[i, 4] = an_dict.get(articul)
                    andrey_changed.append(df.iat[i, column])
                    theOne = theOne.append(df.loc[i], ignore_index=True)
                    photo.append(an_photo.get(articul))
                    newed += 1
                    newed_loc += 1
            except IndexError:
                continue
        df.to_excel(fr'{path}\new\{file}', index=False)
        print(f'Закончил работу с {file} | Обновлено в этой категории: {newed_loc}')
    theOne['ФОТО'] = photo
    theOne = theOne.drop(index=0)

    firstLoad = pd.DataFrame({'Артикул товара': theOne.iloc[0:, 2]})
    firstLoad.to_excel(fr'{path}\new\!1.xlsx', index=False)
    getAllMSC.to_excel(fr'{path}\new\!ALL_IN_ALL', index=False)
    theOne.to_excel(fr'{path}\new\!2.xlsx', index=False)
    print(
        f'Закончил работу. Обновлено файлов артикулов: {newed} | kema: {_ke} | astkol: {_as} | andrey: {_an} | spb: {_id}')
    for i in andrey:
        if i not in andrey_changed:
            print(i)


def startOzon():
    import config.Config as cnf

    inp = input().split()
    marketplace = inp[0].upper()
    id = int(inp[1])


    config = cnf.ConfigInitializer({'marketplace':marketplace,
                                    'id': id})

    return SexOptovik_ozon.SexOptovik_ozon(config)

p = rePattern()
def ExtractPatternFromText(INPUT_DATA, keywords=None):
    if keywords is None:
        keywords = p.ExtractParams
    if not isinstance(INPUT_DATA, list):
        INPUT_DATA = [INPUT_DATA]
    result = {}
    for text in INPUT_DATA:
        for word in keywords:
            pattern = r'{}(?:\s+([\w.-]+(?:\s+[\w.-]+)*))?\s+(\d+[\.,]?\d*)(?:\s*-\s*(\d+[\.,]?\d*))?'.format(word)
            matches = re.findall(pattern, text)
            for match in matches:
                value1 = match[1].replace(',', '.')
                value2 = match[2].replace(',', '.') if match[2] else value1
                value1, value2 = float(value1), float(value2)
                words_between = match[0] if match[0] else ''
                if words_between == '':
                    words_between = 'общий'
                if word not in result:
                    result[word] = {words_between: {'min': value1, 'max': value2}}
                else:
                    result[word] = [result[word], {words_between: {'min': value1, 'max': value2}}]
    return result

def initialize_category(original_categories, description, morph, knn, vectorizer):

    @lru_cache(maxsize=10000)
    def lemmatize(text):
     words = text.split()
     res = list()
     for word in words:
         p = morph.parse(word)[0]
         res.append(p.normal_form)

     return ' '.join(res)
    categories = [lemmatize(cat.lower().replace(">", "").replace("#", "")) for cat in original_categories]

    vectors = vectorizer.fit_transform(categories)

    knn.fit(vectors)

    def get_nearest_categories(desc):
        desc = lemmatize(desc.lower().replace(">", "").replace("#", ""))
        vec = vectorizer.transform([desc])
        distances, indices = knn.kneighbors(vec)
        return [original_categories[i] for i in indices[0]]

    return get_nearest_categories(description)
def test(text, categories):
    morph = pymorphy2.MorphAnalyzer()
    knn = NearestNeighbors(n_neighbors=2, metric='cosine')
    vectorizer = TfidfVectorizer(max_features=1000, ngram_range=(1, 2))
    print(initialize_category(original_categories=categories,
                              description=text,
                              vectorizer=vectorizer,
                              morph=morph,
                              knn=knn))

if __name__ == '__main__':
    # changeBrands(path=r"C:\Users\Anton\Desktop\вб", brand=True)
    # from config.presets.sizes import Sizes
    # A = Sizes().initWeight('вес одного шарика 1.5 кг, вес второго 150 гр, всего 15гр * 10 шт')

    # files = getCategoryFilesPaths(path=r"C:\Users\Anton\Desktop\Folder WB")
    # for i in files:
    #     path = rf"C:\Users\Anton\Desktop\Folder WB\{i}"

    import re

    # text = "длина вибро-втулки 13,5 см, макс. диаметр 3 см; длина втулки 7,5 см, макс. диаметр 2,4 см; длина цепочки 21 см, макс. диаметр 2,4 см"
    # text = "общая длина 24,3 см, длина до пульта 19,5 см, диаметр 3-3,8 см"
    # text = "длина до кольца 17,5 см, вес одного шарика 15г, вес другого 20г"

    # [!]
    # text = "размер трусиков универсальный (40-46), внутренний диаметр колец 3, 4 и 5 см".replace(' и ', '-')

    # text = 'общая длина 12 см, глубина проникновения 9,5 см, длина клит. стим-ра 6,5 см, диаметр 2,1-3 см'
    # texts = [
    #     'длина фаллоса 14 см, макс. диаметр 3 см; длина вагин. фал-са 7 см, макс. диаметр 3,2 см, длина анал. фал-са 9,5 см, макс. диаметр 2,8 см',
    #     'длина до кольца 17,5 см, вес одного шарика 15г, вес другого 20г',
    #     'длина вибро-втулки 13,5 см, макс. диаметр 3 см; длина втулки 7,5 см, макс. диаметр 2,4 см; длина цепочки 21 см, макс. диаметр 2,4 см',
    #     'общая длина 24,3 см, длина до пульта 19,5 см, диаметр 3-3,8 см']
    # print(find_sizes(
    #    'ширина кожаной части оков 3,7 см, длина оков 25,5 см, общая длина сцепки с карабинами 13 см'),
    #    )
    # for k,v in find_sizes(texts).items():
    #     print(k,v)
    startOzon()
