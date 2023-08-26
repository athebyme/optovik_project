from src.AnyOtherCode import main
import openpyxl
import glob
import os

class OzonChangerID():

    seller_code = ''

    def init_articuls(self, paths, seller_code = ''):
        articuls = set()
        errors = set()
        error_xlsx_names = []
        for file in paths:
            isFound = False
            xlsx = openpyxl.load_workbook(file)
            sheets = xlsx.get_sheet_names()
            for i in range(len(sheets)):
                NAME_SHEET = 'Шаблон для поставщика'
                if sheets[i] == NAME_SHEET:
                    xlsx.active = i
                    isFound = True
                    break
            if not isFound:
                error_xlsx_names.append(file[file.rfind("\\")+1:])
                break
            sheet = xlsx.active
            c = 0
            for row in sheet.iter_rows(2, sheet.max_row):
                if row[1].value is None:
                    break
                if c < 2:
                    c += 1
                else:
                    articular = main.Functions.cleanArticul(row[1].value, seller_code=self.seller_code, shortArticular=False)
                    if articular[0]:
                        articuls.add(articular[1])
                    else:
                        errors.add(articular[1])
            if not os.path.exists('../../!changed_oz'):
                os.mkdir('../../!changed_oz')
            if not os.path.exists(f'./!changed_oz/{self.seller_code}'):
                os.mkdir(f'./!changed_oz/{self.seller_code}')
            name_of_xlsx = file[file.rfind("\\")+1:]
            name_of_xlsx = name_of_xlsx[name_of_xlsx.find(' ')+1:]
            path = f'./changed_oz/{self.seller_code}/{name_of_xlsx}'
            create_parent_xlsx = openpyxl.Workbook()
            create_parent_xlsx.save(path)

            _parent_xlsx = openpyxl.load_workbook(path)
            _parent_xlsx._add_sheet(file)
            _parent_xlsx.save(path)
        print(f'Ошибочных таблиц: {len(error_xlsx_names)}\n')
        if len(error_xlsx_names)!=0:
            print(f'Ошибочные таблицы: {error_xlsx_names}\n')
        return articuls, errors

    def start(self):
        path = main.Functions.getFolderFile(1, item='файлов категорий c озона')
        all_paths = glob.glob(rf"{path}/*.xlsx")
        #print(all_paths)
        t = len(all_paths) % 10
        name = ''
        if t == 1:
            name = 'файл'
        elif t == 0 or 5 <= len(all_paths) < 21:
            name = 'файлов'
        elif 2 <= t <= 3:
            name = 'файла'
        else:
            name = 'файлов'
        print(f'Обнаружено {len(all_paths)} {name}\nПодсчет артикулов.')
        articuls = self.init_articuls(all_paths)
        print(f'\nОбнаружено {len(articuls[0])} артикулов')
        input('DEBUG')

    def __init__(self, seller_code):
        print('\nВы запустили скрипт смены артикула поставщика для озона.\n')
        self.seller_code = str(seller_code)
        # articuls, errors = self.init_articuls(path = )
        # print(articuls, errors)

