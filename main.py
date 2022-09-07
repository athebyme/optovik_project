import codecs
import socket
from dataclasses import dataclass
from pathlib import Path

import google.auth.exceptions
import google_auth_oauthlib
import httplib2.error
from bs4 import BeautifulSoup
from selenium import webdriver
from tkinter import filedialog
from datetime import date
from pyfiglet import Figlet

import SexOptovik_wb
from Google import Create_Service
from googleapiclient.http import MediaIoBaseDownload
from itertools import product

import openpyxl
import re
import math
import os
import random
import sys
import time
import tkinter as tk
import urllib, requests
import pandas as pd
import shutil
import io

all_providers = [
    'astkol',  # 0
    'kema',  # 1
    'andrey'  # 2
    'sex_optovik'
]


@dataclass
class MainData():
    provider: str
    seller_code: str
    excelFiles_path: str
    imgsFiles_path: str
    barcodeFile_path: str


class Functions:
    provider = ''
    seller_code = ''
    supported_providers = [
        'as_price_and_num_rrc.csv',
        'ke_price_and_num_rrc.csv'
    ]
    paths = []
    abs_path = os.getcwd()

    def __int__(self, abs_path, sellerCode, provider):
        self.provider = self.getProvider()
        self.seller_code = self.getSellerCode()
        self.paths = []

    @staticmethod
    def uploadBarcodes(path, SET_OF_EXIST_BARCODES=set()):
        if path == '': return False
        barcodes = SET_OF_EXIST_BARCODES.copy()
        try:
            with open(path, encoding='utf8') as f:
                t = f.readline()
                for lines in f:
                    barcodes.add(int(lines))
            print('Штрихкоды успешно загружены')
            return barcodes
        except UnicodeError:
            print('Ошибка кодирования файла. Попробуйте снова')
            sys.exit(0)
        except BaseException:
            print('Ошибка.')
            sys.exit(0)

    @staticmethod
    def getSellerCode():
        time.sleep(0.3)
        seller_code = input('Введите код продавца: ')
        return seller_code

    @staticmethod
    def getProvider():
        time.sleep(0.3)
        provider = input('Введите название поставщика: ')
        return provider

    @staticmethod
    def getPath(type):
        time.sleep(1)
        root = tk.Tk()
        root.withdraw()
        if type == 1:
            return filedialog.askdirectory()
        else:
            return filedialog.askopenfilename()

    @staticmethod
    def getFolderFile(type, item):
        time.sleep(0.3)
        try:
            if type == 1:
                print(f'Выберите папку, куда будут загружаться {item}')
            else:
                print(f'Выберите файл, откуда будет(ут) выбрано(ы) {item}')
            return Functions.getPath(type)
        except OSError:
            return False

    @staticmethod
    def wildberries_site():
        driver = webdriver.Chrome("E:\chromedriver_win32\chromedriver.exe")
        # url =
        driver.get('https://seller.wildberries.ru')
        print(driver)

    @staticmethod
    def clean_for_data(articular, seller_code):
        temp = ''.join(articular)
        articular = ''.join(articular)
        # set_of_post_codes = [
        #     'Z1C1A0t-', 'W1C1Ayt', 'Z1C1L', 'Z1C1K', 'W1C1K', 'W1C1V', 'Z1C1A', 'W1C1L', 'W1C1A', 'Z1C1V'
        # ]
        set_of_post_codes = [
            'UNCOMMENT PREV ARRAY IF CHECK Z1 NEEDED'
        ]
        check_any_code_in = []
        for i in set_of_post_codes:
            check_any_code_in.append(i in articular)

        if 'id' in articular:
            articular = articular[articular.find('-') + 1:]
            articular = articular[:articular.find('-')]
            if 4 + len(seller_code) + len(articular) != len(temp):
                return False, temp
        elif any(check_any_code_in):
            for i in set_of_post_codes:
                if i in articular:
                    current_code = i
                    break
            ind = articular.find(current_code)
            if len(articular[:ind]) != len(seller_code):
                seller_code = articular[:ind]
            articular = articular[ind:]

            c = len(current_code)
            articular = articular.replace(current_code, '')

            if len(articular) + c + len(seller_code) != len(temp):
                return False, temp
        else:
            return False, temp
        return True, articular

    def getData(self, path, seller_code, marketplace = 'wb'):
        articuls = set()
        errors = set()
        enc = ''
        try:
            with open(path, 'r') as f:
                t = f.readline()
                for lines in f:
                    if marketplace == 'wb':
                        articular = self.clean_for_data(list(map(str, lines.split(';')))[3], seller_code)
                    elif marketplace == 'oz':
                        articular = self.clean_for_data(list(map(str, lines.split(';')[0])), seller_code)
                    if articular[0]:
                        articuls.add(articular[1])
                    else:
                        errors.add(articular[1])
            print('Существующие товары успешно загружены')
            return articuls, errors


        # TypeError

        except BaseException:
            xlsx = openpyxl.load_workbook(path)
            sheet = xlsx.active
            for row in sheet.iter_rows(2, sheet.max_row):
                articular = self.clean_for_data(row[3].value, seller_code=seller_code)
                if articular[0]:
                    articuls.add(articular[1])
                else:
                    errors.add(articular[1])
            return articuls, errors

    @staticmethod
    def download_universal(url='', path_def=''):
        try:
            print(f'\n\nПытаюсь загрузить {url}\nПуть: {path_def}')
            path = f'{path_def}/{url[url.rfind("/") + 1:]}'
            excel_file = open(path, 'wb')
            excel_file.write(urllib.request.urlopen(url).read())
            excel_file.close()
        except OSError:
            input(f'\nЗакройте файл {url[url.rfind("/") + 1:]}. И нажмите enter\n ')
            print(f'\n\nПытаюсь загрузить {url}\nПуть: {path_def}')
            path = f'{path_def}/{url[url.rfind("/") + 1:]}'
            try:
                excel_file = open(path, 'wb')
                excel_file.write(urllib.request.urlopen(url).read())
                excel_file.close()
                return path
            except OSError:
                print('Непредвиденная ошибка')
                sys.exit(0)
        else:
            print(f'\n\nНовые товары {url} успешно загружены.\nПуть: {path}\n')
            return path

    def downloadNewGoods(self, provider, abs_path):

        # 0 = astkol
        # 1 = kema

        if provider == -1:
            name = ['as_price_and_num_rrc.csv', 'ke_price_and_num_rrc.csv']
            url = ['http://www.sexoptovik.ru/mp/as_price_and_num_rrc.csv',
                   'http://www.sexoptovik.ru/mp/ke_price_and_num_rrc.csv']
        if provider == 0:
            name = ['as_price_and_num_rrc.csv']
            url = ['http://www.sexoptovik.ru/mp/as_price_and_num_rrc.csv']
        elif provider == 1:
            name = ['ke_price_and_num_rrc.csv']
            url = ['http://www.sexoptovik.ru/mp/ke_price_and_num_rrc.csv']
        print('\nСкачиваю файл с новыми товарами\n')

        # for i in range(len(name)):
        for it in name:
            path = Path(abs_path, it)
            try:
                os.remove(path)
            except OSError:
                print('Закройте файл(ы) с товарами')
            item = self.download_universal(it, path)

        #    if name[i] == name[len(name) - 1]:
        print(f'Новые товары {name} успешно загружены.\n')
        return 1

    def workWithExcel(self, code_provider):
        curw = os.getcwd()
        self.downloadNewGoods(code_provider, curw)
        paths = []
        if code_provider == -1:
            for name in self.supported_providers:
                paths.append(Path(curw, name))
        if code_provider == 0:
            paths.append(Path(curw, self.supported_providers[0]))
        if code_provider == 1:
            paths.append(Path(curw, self.supported_providers[1]))

        self.paths = paths.copy()

        for p in paths:
            self.openExcelFile(p)

    #

    @staticmethod
    def openExcelFile(path):
        with open(path, 'r', encoding='utf8') as excel_file:
            f = excel_file.readline()
            while '"Раздел";"id";"ГруппаТоваров";"Артикул";"Размер";' not in f:
                f = excel_file.readline()

    @staticmethod
    def category(text, extra):
        cats_ex = {'набор': 'Наборы игрушек для взрослых', 'анальн шарик': 'Анальные шарики', 'плаг': 'Анальные пробки',
                   'анал плаг': 'Анальные пробки', 'анальные бусы': 'Анальные бусы',
                   'анальная пробка': 'Анальные пробки', 'анальн фаллоимитатор': 'Фаллоимитаторы',
                   'анальные расширители': 'Расширители гинекологические', 'анальны вибромассажер': 'Анальные бусы',
                   'анальный душ': 'Фаллоимитаторы', 'массажер простат': 'Массажеры простаты',
                   'эрекционн кольц': 'Эрекционные кольца', 'вагинальн шарик': 'Вагинальные шарики',
                   'анальн пробк': 'Анальные пробки', 'мастурбатор': 'Мастурбаторы мужские',
                   'зажим для сосков': 'Зажимы для сосков', 'зажим на соски': 'Зажимы для сосков',
                   'браслет': 'Браслеты эротик', 'кляп': 'Кляпы эротик', 'многохвостн плет': 'Плетки эротик',
                   'пояс верност': 'Пояса верности', 'ошейник': 'Ошейники эротик',
                   'расширител влагальщн': 'Расширители гинекологические', 'сбру': 'Боди эротик',
                   'наручники': 'Наручники эротик', 'маск': 'Маски эротик', 'стэк': 'Стэки эротик',
                   'фиксаци': 'Фиксаторы эротик', 'чокер': 'Чокеры эротик', 'шлепалк': 'Шлепалки эротик',
                   'вибропул': 'Вибропули', 'свеч': 'Свечи эротик', 'свеча': 'Свечи эротик',
                   'массажн': 'Массажные средства эротик',
                   'очищающ средств для игрушек': 'Средства для очистки секс-игрушек',
                   'парфюмерия': 'Ароматизаторы эротик', 'шарик': 'Анальные шарики',
                   'виброяйцо': 'Виброяйца', 'вакуумны стимулятор': 'Вакуумно-волновые стимуляторы',
                   'вакуумные': 'Вакуумные помпы эротик', 'плать': 'Платья эротик', 'трус': 'Трусы эротик',
                   'чулки': 'Чулки эротик'}
        cats = {"фаллоимитатор": "Фаллоимитаторы", "лубрикант": "Лубриканты",
                "стимулятор клитора": "Вагинальные тренажеры", "бдсм": "Комплекты БДСМ",
                "вагин": "Мастурбаторы мужские", "вибромассажер": "Вибраторы",
                "игрушки из стекла": "Наборы игрушек для взрослых", "интимная косметика": "Уходовые средства эротик",
                'комплекты белья': 'Комплеты эротик', "куклы": "Секс куклы", "менструальные чаши": "Спринцовки эротик",
                "насадки и кольца": "Эрекционные кольцо", "помпы": "Гидропомпы эротик",
                "продукция с  феромонами": "Концентраты феромонов", "секс-машин": "Секс машины"
            , "скоро в продаже": "Наборы игрушек для взрослых",
                "стимуляторы итора": "Вибраторы", "страпон": "Страпоны",
                "электростимулятор": "Электростимуляторы", "эльмято": "Наборы игрушек для взрослых",
                "эротическое елье": "Комплекты БДСМ",
                "промо": "Наборы игрушек для взрослых", "анальные стимуляторы": "Вибраторы",
                "игрушка из стекла": "Сувениры эротик",
                "наборы секс-игрушек": "Наборы игрушек для взрослых", "плаг": "Анальные пробки",
                'стимулятор клитор': 'Вибраторы'}
        t = text.lower()
        _extra = extra
        if True:
            for i in cats_ex.keys():
                if (i in t): return cats_ex.get(i)
        for i in cats.keys():
            if i in _extra: return cats.get(i)
        if 'боди' in t or 'футболк' in t:
            return 'Боди эротик'
        elif 'трус' in t:
            return 'Трусы эротик'
        elif 'комплект' in t:
            return 'Наборы игрушек для взрослых'
        elif 'ремень' in t:
            return 'Комплекты БДСМ'
        elif 'на фаллос' in t:
            return 'Насадки на член'
        elif 'насадки без вибрации' in t:
            return 'Насадки на член'
        elif 'насадка на пенис' in t:
            return 'Насадки на член'
        elif 'насадка' in t:
            return 'Насадки на страпон'
        elif 'качели' in t:
            return 'Секс качели'
        elif 'кресл' in t:
            return 'Секс кресла'
        elif 'фиксац' in t:
            return 'Фиксаторы эротик'
        elif 'набор' in t:
            return 'Наборы игрушек для взрослых'
        elif 'простын' in t:
            return 'Простыни БДСМ'
        elif 'пилон' in t:
            return 'Комплекты БДСМ'
        else:
            return 'Наборы игрушек для взрослых'

    @staticmethod
    def barcodes(SET_OF_BARCODES):
        SET_OF_NEW_BARCODES = set()

        return barcodes

    @staticmethod
    def getSex(data):
        if data.count('муж') != 0:
            return 'Мужской'
        elif data.count('жен') != 0:
            return 'Женский'
        return 'Женский'

    @staticmethod
    def checkClothes(text):
        t = text.lower()
        if ('боди' in t) or ('футболк' in t) or ('трус' in t) or ('чулк' in t) or ('плать' in t):
            return '42-48'
        else:
            return ''

    @staticmethod
    def setBarcode(barcodes):
        barcode = barcodes[0]
        barcodes.pop(0)
        return barcode, barcodes

    @staticmethod
    def find_material(text):
        t = text
        if ('Материал' in t):
            ind = t.rfind('Материал') + 10
            t = t[ind:]
            t = t[:t.find('<')]
        else:
            t = 'Пластик'
        return t

    def xlsx_photo_wb(self, dict_art_ph, seller_code='', path='', text='', name = 'file_photo_wb'):
        xlsx = {1: ["Артикул товара"], 2: ["Медиафайлы"]}
        for art,ph in dict_art_ph.items():
            xlsx[1].append(art)
            xlsx[2].append(ph)
        Functions.save_data(self, data=xlsx, path=path, seller_code=seller_code, original_name=name,
                            text_print=text)

    @staticmethod
    def cleanText(text):
        t = BeautifulSoup(text, "lxml").text.replace("\xc2\xa0", " ").replace('&nbsp', '').replace('&quot', '')
        if len(t) > 5000:
            t = t[:5000]
            t = t[:t.rfind('.')]
        t = t.replace('°', '')
        t = t.replace("\\", '')
        t = t.replace('–', '')
        t = t.replace('’', '')
        t = t.replace('…', '.')
        t = t.replace('x000D', '')
        t = t.replace('¶', '')
        t = t.replace('«', '"')
        t = t.replace('»', '"')
        t = t.replace('®', '')
        t = t.replace('`', '"')
        t = t.replace('_', '')
        t = t.replace('”', '"')
        t = t.replace('“', '"')
        t = t.replace('™', '')
        t = t.replace('=', '-')
        t = t.replace('obj', '')
        t = t.replace('—', '-')
        t = t.replace('•', '')
        t = t.replace('–', '-')
        t = t.replace('	', ' ')
        t = t.replace('´', '-')
        t = t.replace('/', ';')
        t = t.replace('<', ' менее ')
        t = t.replace('>', ' более ')
        t = t.replace('í', 'i')
        t = t.replace('à', 'a')
        t = t.replace('é', 'e')
        t = t.replace('±', '+-')
        t = t.replace('*', '')
        t = t.replace('	', ' ')
        t = t.replace("&quot;", '')
        t = t.replace("&quot", '')
        t = t.replace("×", 'x')
        t = t.replace("а́", 'а')
        return t

    @staticmethod
    def weight(text):
        t = text
        if 'Вес' in t:
            z = len(t) // 2
            if t.rfind('Вес') > z:
                t = t[z:]
            else:
                t = t[:z]
            ind = t.rfind('Вес') + 4
            t = t[ind:]
            if t.find('<') < t.find('.'):
                ind = t.find('<')
            else:
                ind = t.find('.')
            t = t[:ind]
            t = re.sub("[^0-9]", "", t)
            return [float(t), round(float(t) / 1000, 2)]
        else:
            return ['На упаковке', 'На упаковке']

    @staticmethod
    def weight_bad(text):
        t = text
        if 'Вес' in t:
            ind = t.find('Вес') + 3
            t = t[ind:]
            ind = t.find('.')
            if ind > 40:
                ind = t.find(',')
            if ind == -1:
                if 'Вес' in t:
                    t = t[t.find('Вес') + 3:]
                ind = t.find('кг')
                ed = 1
                if ind == -1:
                    ind = t.find('гр')
                    ed = 0
                t = t[:ind]
                t = t[t.find(':') + 2:]
                t = t.replace(',', '.')
                if "\t" in t:
                    t = t[t.find('\t') + 1:]
                if ed:
                    return [float(t) * 1000, float(t)]
                else:
                    return [float(t), round(float(t) / 1000, 2)]
            t = t[:ind - 2]
            t = t[t.find(':') + 2:]
            try:
                t = t.replace(',', '.')
                return [float(t), round(float(t) / 1000, 2)]
            except ValueError:
                if 'гр' in t:
                    t = t[:t.find('гр')]
                else:
                    if '(' in t:
                        t = t[:t.find('(')]
                    else:
                        t = t[-1:]
                try:
                    t = t.strip()
                    return [float(t), round(float(t) / 1000, 2)]
                except ValueError:
                    # t = t[:-1]
                    ind = 0
                    for z in t:
                        if not (z.isdigit()):
                            ind += 1
                    t = t[ind - 1:]
                    if 'см' in t:
                        return ['На упаковке', 'На упаковке']
                    try:
                        return [float(t), round(float(t) / 1000, 2)]
                    except ValueError:
                        t = text
                        ind = t.rfind('Вес') + 3
                        t = t[ind:]
                        ind = t.find('.')
                        t = t[:ind]
                        ind = t.find(':')
                        if ind > 15 or ind == -1:
                            t = t[t.find('\t'):]
                        else:
                            t = t[ind + 2:]
                        t = t[:-2]
                        t = t.replace(',', '.')
                        t = t.replace(' ', '')
                        return [float(t.strip()), round(float(t.strip()) / 1000, 2)]
        return ['На упаковке', 'На упаковке']

    @staticmethod
    def getSizes(text):
        t = text
        ed = 'см'
        if 'упаковки' in t:
            ind = t.rfind('упаковки')
            temp = t[ind - 5:]
            temp = temp[:temp.find('<')]
            if 'мер' in temp:
                ind = temp.rfind('упаковки') + 9
                temp = temp[ind:]
                if temp.rfind('.') != -1:
                    temp = temp[:temp.rfind('.')]
                temp = temp.replace('х', 'x')
                temp = re.sub("[^0-9,x .]", "", temp)
                sizes_array = list(map(lambda temp: temp.strip(), temp.split('x')))
                return sizes_array
        lenth = height = width = 'Не указано'
        if 'Длина' in t:
            temp = text
            lenth = temp[temp.rfind('Длина') + 6:]
            ind = lenth.find('<')
            if ind == -1:
                ind = lenth.rfind('.')
            lenth = lenth[:ind]
            lenth = re.sub("[^0-9]", "", lenth)
        if 'Ширина' in t:
            temp = text
            width = temp[temp.rfind('Ширина') + 7:]
            ind = width.find('<')
            if ind == -1:
                ind = width.rfind('.')
            width = width[:ind]
            width = re.sub("[^0-9]", "", width)
        if 'Высота' in t:
            temp = text
            height = temp[temp.rfind('Высота') + 7:]
            ind = height.find('<')
            if ind == -1:
                ind = height.rfind('.')
            height = height[:ind]
            height = re.sub("[^0-9]", "", height)
        return [width, lenth, height]

    @staticmethod
    def Osob(data):
        if 'скорост' in data.lower():
            ind = data.find('скорост') - 70
            t = data[ind:]
            ind = data.find('.') + 2
            t = t[ind:]
            ind = data.find('.')
            t = t[:ind]
            return t
        else:
            return 'Смотрите в описании'

    @staticmethod
    def getCountry(text):
        t = text
        ind = t.rfind('Производитель')
        if ind != -1:
            t = t[ind + 15:]
            t = t[t.rfind(',') + 1:].strip()
            return t
        else:
            return 'Китай'

    def save_data(self, data, seller_code, path=os.getcwd(), original_name=str(round(time.time())), _print=True,
                  _full=False, text_print = ''):
        file = Path(path, f'{seller_code}_{original_name}.xlsx')
        try:
            os.remove(file)
        except BaseException:
            pass
        #file = f'./!parsed_full/{seller_code}_{original_name}.xlsx'
        # if _full:
        #     try:
        #         shutil.rmtree(f'./!parsed_full/{seller_code}/')
        #     except OSError:
        #         # print('Пожалуйста, закройте файл с товарами')
        #         pass
        success = False

        while not success:
            try:
                with open(file, 'w'):
                    print('')
                success = True
            except PermissionError:
                input('Пожалуйста, закройте файл с товарами и нажмите Enter')
        success = False
        x = pd.DataFrame(data)
        while not success:
            try:
                x.to_excel(file, index=False, header=False)
                if _print: print(f'Успешно {text_print}!')
                success = True
            except PermissionError:
                input('Закройте файл, нажмите Enter')

        time.sleep(0.5)

    @staticmethod
    def downloadMedia(media):
        directory_media = Path(os.getcwd(), 'media')
        try:
            shutil.rmtree(directory_media)
            os.rmdir(directory_media)
        except OSError:
            os.mkdir(directory_media)
        c = 0
        l = len(media)
        for k, v in media.items():
            ind = 1
            barcode = k
            path = Path(directory_media, str(barcode))
            os.mkdir(path)
            urls = v.copy()
            for url in urls:
                img = open(Path(path, str(ind) + '.jpg'), 'wb')
                try:
                    img.write(urllib.request.urlopen(url).read())
                except BaseException:
                    time.sleep(0.1)
                ind += 1
                img.close()
            c += 1
            print(f'Загружено фотографии для {c} товаров из {l}. Выполнено на {round(c / l * 100, 2)} %')
        return directory_media

    def get_wb_cats(self, cwd=os.getcwd()):
        file = self.getFolderFile(0, ' категории с Wildberries')
        # file = "C:\Users\Антон\Desktop\WB_cats.docx"
        set_of_articulars = set()
        try:
            with open(file) as f:
                for line in f:
                    set_of_articulars.add(line.rstrip())
            path = Path(cwd, 'pool', 'SexOptovik', 'wb_cats.txt')
        except UnicodeDecodeError:
            # doc = docx.getdocumenttext(file)
            text = []
        try:
            set_of_articulars.remove('')
        except KeyError:
            print('Вы выбрали неподходящий файл. Выберите верные категории\nПопробуйте еще раз')
            sys.exit(0)
        if not Path.exists(path):
            with open(path, 'w+') as f:
                for items in set_of_articulars:
                    f.write(items + '\n')
        else:
            CHOOSE = input('СОХРАНИТЬ ИЗМЕНЕНИЯ КАТЕГОРИЙ WILDBERRIES ?\n1 = да\n0 = нет\n')
            while not CHOOSE.isdigit() or CHOOSE not in '10':
                CHOOSE = input('ПОЖАЛУЙСТА, ВВЕДИТЕ ЧИСЛО 1 ИЛИ 0: ')
            if CHOOSE == '1':
                os.remove(path)
                with open(path, 'w+') as f:
                    for items in set_of_articulars:
                        f.write(items + '\n')
        return set_of_articulars

    @staticmethod
    def showText(text, font='big'):
        try:
            # univers #basic #big #doom
            view_text = Figlet(font=font)
            print(view_text.renderText(text))
        except BaseException:
            print(text)

    @staticmethod
    def google_driver(google_ids=[], file_names=[], path_os_type='./'):
        _COUNT_TRIES_BEFORE_EXIT = 2
        print('')
        COUNT_DOWLOADED = 0
        CLIENT_SECRET_FILE = '.\client_secrets.json'
        API_NAME = 'drive'
        API_VERSION = 'v3'
        SCOPES = ['https://www.googleapis.com/auth/drive']

        success = False
        _try = 1
        while not success:
            try:
                service = Create_Service(CLIENT_SECRET_FILE, API_NAME, API_VERSION, SCOPES, )
                for google_id, file_name in zip(google_ids, file_names):
                    request = service.files().get_media(fileId=google_id)

                    fh = io.BytesIO()
                    downloader = MediaIoBaseDownload(fd=fh, request=request)

                    done = False

                    while not done:
                        status, done = downloader.next_chunk()
                        print(f"Загружено с диска {file_name}: %d%%." % int(status.progress() * 100))
                        COUNT_DOWLOADED +=1
                    fh.seek(0)

                    with open(os.path.join(path_os_type, file_name), 'wb') as f:
                        f.write(fh.read())
                        f.close()
                success = True
            except OSError as ex:
                input(f'Ошибка! {ex}.\nПроверьте соединение и нажмите любую клавишу.')
                if _try > _COUNT_TRIES_BEFORE_EXIT:
                    _try += 1
                else:
                    print(f'Это уже {_try} попытка. \n'
                          f'Попробуйте проверить данные и перезапустить программу)')
                    sys.exit(0)
            except httplib2.error.ServerNotFoundError:
                input('Проверьте соединение и нажмите любую клавишу')
            except google.auth.exceptions.RefreshError:
                print('Токен устарел. Необходимо произвести замену токена.')
                if os.path.isfile('./token_drive_v3.pickle'):
                    os.remove('./token_drive_v3.pickle')
                    print('Устаревший токен успешно удален. Необходимо пройти авторизацию заново.')
                    time.sleep(3)
                else:
                    print('Файл токена не найден в текущем местоположении. Выберите его самостоятельно')
                    path_token = Functions.getFolderFile(0, item=' файл токена google')
                    os.remove(path_token)
                    # return
            print(f'Успешно загружено {COUNT_DOWLOADED} / {len(google_ids)} файлов\n')
            time.sleep(1.5)

    def uploadFromFile(self, file_path, seller_code='', isSet=True):
        if isSet:
            set_res = set()
            first = True
            with open(file_path) as f:
                for line in f:
                    if first:
                        set_res.add(line[3:])
                        first = False
                    else:
                        set_res.add(line)
            return set_res
        else:
            dict_res = {}
            with open(file_path) as f:
                success = False
                while not success:
                    try:
                        three_check = 0
                        for line in f:
                            res_d = ''
                            _id = line[:line.find(';')]
                            t_l = self.cleanText(line[line.find(';') + 1:]).replace('"', '')
                            dict_res.setdefault(_id, t_l)
                            success = True
                        return dict_res
                    except PermissionError:
                        input('Закройте файлы XLSX и нажмите любую кнопку.')

    @staticmethod
    def for_sizes_parse(array_4XL):
        size = ['XS', 'S', 'M', 'L', 'XL', 'XXL', 'XXXL', '4XL']
        size_ru = array_4XL.copy()
        dict = {}
        for i in range(len(size)):
            dict.setdefault(size[i], size_ru[i])
        return dict


class kema_parser(Functions):
    provider = 'kema'
    code = 'W1C1K'
    provider_code = ''
    barcodes = []
    ALL_MEDIA = {}
    EXIST_DATA = set()

    wb_pattern = {1: ['Номер карточки'], 2: ['Категория'], 3: ['Цвет'], 4: ['Бренд'], 5: ['Пол'],
                  6: ['Название'], 7: ['Артикул товара'], 8: ['Размер'],
                  9: ['Рос. размер'], 10: ['Баркод товара'], 11: ['Цена'], 12: ['Состав'], 13: ['Описание'],
                  14: ['Вес (г)'], 15: ['Вес без упаковки (кг)'], 16: ['Вес с упаковкой (кг)'],
                  17: ['Вес товара без упаковки (г)'],
                  18: ['Вес товара с упаковкой (г) (г)'], 19: ['Вид вибратора'], 20: ['Вид лубриканта'],
                  21: ['Вид мастурбатора'],
                  22: ['Вкус'], 23: ['Возрастные ограничения'],
                  24: ['Высота предмета (см)'], 25: ['Высота упаковки (см)'], 26: ['Глубина предмета'],
                  27: ['Глубина упаковки (см)'], 28: ['Диаметр предмета (см)'], 29: ['Действие лубриканта'],
                  30: ['Диаметр секс игрушки'], 31: ['Количество предметов в упаковке'],
                  32: ['Количество предметов в упаковке (шт.)'], 33: ['Комплектация'],
                  34: ['Материал изделия'], 35: ['Модель'],
                  36: ['Объем'], 37: ['Объем средства'], 38: ['Объем товара'], 39: ['Особенности белья'],
                  40: ['Особенности продукта'], 41: ['Особенности секс игрушки'],
                  42: ['Рабочая длина секс игрушки'], 43: ['Питание'],
                  44: ['Упаковка'],
                  45: ['Ширина предмета'], 46: ['Ширина упаковки (см)'], 47: ['Медиафайлы'], 48: ['Страна производства']
                  }

    parsed_kema = wb_pattern.copy()

    def __init__(self, provider_code, barcodes, data):
        self.provider = 'kema'
        self.provider_code = str(provider_code)
        self.barcodes = barcodes.copy()
        self.seller_code = Functions.getSellerCode()

        # добавление базы существующих артикулов
        for items in data:
            self.EXIST_DATA.add(items)

    def start_parser(self):
        all_media = {}

        path = Path(Functions.abs_path, 'ke_price_and_num_rrc.csv')
        with open(path, 'r', encoding='utf8') as excel_file:
            f = excel_file.readline()
            while '"Раздел";"id";"ГруппаТоваров";"Артикул";"Размер";' not in f:
                f = excel_file.readline()
            it = 0
            for line in excel_file:
                DATA = list(map(lambda line: line.replace('"', ''), line.split(';')))
                if DATA[1] not in self.EXIST_DATA:
                    opis = DATA[7]

                    cat = Functions.category(opis, DATA[27])
                    razm = Functions.checkClothes(opis)
                    barcode, self.barcodes = Functions.setBarcode(self.barcodes)
                    material = Functions.find_material(opis)
                    cleantext = Functions.cleanText(opis)
                    weight_g, weight_kg = Functions.weight(opis)
                    try:
                        weight_pack_g, weight_pack_kg = weight_g + 10, weight_kg + 0.015
                    except TypeError:
                        weight_pack_g, weight_pack_kg = '+-10 гр.', '+-0.01 кг.'
                    if 'Вибр' in cat or 'Вибр' in cleantext:
                        vibrator_type = DATA[6]
                    else:
                        vibrator_type = '--Не является вибратором--'
                    if 'Мастур' in cat:
                        masturbator_type = DATA[6]
                    else:
                        masturbator_type = '--Не является мастурбатором--'
                    if 'Лубри' in cat:
                        lubricant_type = DATA[6]
                        deist_lbr = DATA[6]
                    else:
                        deist_lbr = lubricant_type = '--Не является лубрикантом--'

                    if it == 10:
                        print(' ')

                    width, length, height = Functions.getSizes(opis)

                    try:
                        try:
                            t = float(width)
                        except ValueError:
                            t = ''
                        width_item = t - round(random.uniform(0.1, 0.3), 2)
                    except TypeError:
                        width_item = 0
                    try:
                        try:
                            t = float(length)
                        except ValueError:
                            t = ''
                        length_item = t - round(random.uniform(0.1, 0.3), 2)
                    except TypeError:
                        length_item = 0
                    try:
                        try:
                            t = float(height)
                        except ValueError:
                            t = ''
                        height_item = t - round(random.uniform(0.1, 0.3), 2)
                    except TypeError:
                        height_item = 0
                    try:
                        try:
                            width = float(width)
                            height = float(width)
                        except ValueError:
                            diameter = 0
                        diameter = round(math.sqrt(width ** 2 + height ** 2))
                    except TypeError:
                        diameter = 0
                    try:
                        V = str(round(height_item * width_item * length_item, 2)) + ' см^3'
                    except TypeError:
                        V = 0
                    media = []
                    for i in range(19, 27):
                        if DATA[i] != '':
                            media.append(DATA[i])
                    media.sort()

                    articul = self.provider_code + self.code + DATA[1]

                    kema_local = {2: cat, 3: DATA[17], 4: DATA[18], 5: Functions.getSex(opis), 6: DATA[6],
                                  7: articul, 8: razm, 9: razm, 10: barcode, 11: 99999, 12: DATA[6]
                        , 13: cleantext, 14: weight_g, 15: weight_kg, 16: weight_pack_kg, 17: weight_g,
                                  18: weight_pack_g,
                                  19: vibrator_type, 20: lubricant_type, 21: masturbator_type,
                                  22: 'На вкус как ' + material, 23: '18+', 24: height_item, 25: height,
                                  26: height_item,
                                  27: height, 28: diameter, 29: deist_lbr, 30: diameter, 31: 1, 32: '1 шт.',
                                  33: '1 шт.', 34: material, 35: DATA[2], 36: 'Смотрите в описании',
                                  37: 'Смотрите в описании',
                                  38: V, 39: DATA[6], 40: DATA[6], 41: DATA[16] + '. ' + Functions.Osob(cleantext),
                                  42: height_item,
                                  43: 'Смотрите в описании', 44: 'Непрозрачная анонимная упаковка', 45: width_item,
                                  46: width, 47: media[0], 48: Functions.getCountry(opis)}
                    it += 1
                    print(it, kema_local)
                    for k, v in kema_local.items():
                        self.parsed_kema[k].append(v)
                    self.ALL_MEDIA.setdefault(articul, [])
                    for urls in media:
                        self.ALL_MEDIA[articul].append(urls)
            for i in range(1, len(self.parsed_kema[7])):
                self.parsed_kema[1].append(i)
            Functions.save_data(self, self.parsed_kema)
        return self.ALL_MEDIA


class Andrey(Functions):
    def __init__(self):
        pass

class SexOptovik(Functions):
    cwd = os.getcwd()
    optovik_items = {}
    current_cats_wb = set()
    seller_code = ''
    size_img = 650
    CONST_AMOUNT_OF_XLSX_ITEMS = 10

    parsed_items = {1: ['Номер карточки'], 2: ['Категория'], 3: ['Цвет'], 4: ['Бренд'], 5: ['Пол'], 6: ['Название'],
                    7: ['Артикул товара'], 8: ['Размер'], 9: ['Рос. размер'], 10: ['Баркод товара'],
                    11: ['Цена'], 12: ['Состав'], 13: ['Медиафайлы'], 14: ['Описание'],
                    15: ['Страна производства'],
                    16: ['Особенности секс игрушки'], 17: ['Особенности модели'], 18: ['Материал'],
                    19: ['Наличие батареек в комплекте'], 20: ['Объем'], 21: ['Объем (мл)'],
                    22: ['Объем средства'], 23: ['Ширина предмета'], 24: ['Ширина упаковки'], 25: ['Длина (см)'],
                    26: ['Длина секс игрушки'],
                    27: ['Рабочая длина секс игрушки'], 28: ['Высота предмета'], 29: ['Высота упаковки'],
                    30: ['Глубина предмета'], 31: ['Глубина упаковки'], 32: ['Диаметр'],
                    33: ['Диаметр секс игрушки'], 34: ['Вид вибратора'], 35: ['Вес без упаковки'], 36: ['Вес(г)'],
                    37: ['Вес средства'], 38: ['Вес товара без упаковки(г)'], 39: ['Вес товара с упаковкой(г)'],
                    40: ['Комплектация'], 41: ['Количество предметов в упаковке'], 42: ['Упаковка']

                    }
    parsed_items_100_items = {1: ['Номер карточки'], 2: ['Категория'], 3: ['Цвет'], 4: ['Бренд'], 5: ['Пол'],
                              6: ['Название'],
                              7: ['Артикул товара'], 8: ['Размер'], 9: ['Рос. размер'], 10: ['Баркод товара'],
                              11: ['Цена'], 12: ['Состав'], 13: ['Медиафайлы'], 14: ['Описание'],
                              15: ['Страна производства'],
                              16: ['Особенности секс игрушки'], 17: ['Особенности модели'], 18: ['Материал'],
                              19: ['Наличие батареек в комплекте'], 20: ['Объем'], 21: ['Объем (мл)'],
                              22: ['Объем средства'], 23: ['Ширина предмета'], 24: ['Ширина упаковки'],
                              25: ['Длина (см)'],
                              26: ['Длина секс игрушки'],
                              27: ['Рабочая длина секс игрушки'], 28: ['Высота предмета'], 29: ['Высота упаковки'],
                              30: ['Глубина предмета'], 31: ['Глубина упаковки'], 32: ['Диаметр'],
                              33: ['Диаметр секс игрушки'], 34: ['Вид вибратора'], 35: ['Вес без упаковки'],
                              36: ['Вес(г)'],
                              37: ['Вес средства'], 38: ['Вес товара без упаковки(г)'],
                              39: ['Вес товара с упаковкой(г)'],
                              40: ['Комплектация'], 41: ['Количество предметов в упаковке'], 42: ['Упаковка']

                              }

    @staticmethod
    def parse_sizes(text, type):
        eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
        eds_en = ['length', 'width', 'height', 'diameter', 'depth']
        if ', ' not in text or type == 'clothe':
            t = text
            data = []
            eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина', 'вес']
            indx = []
            indx_first = []
            data = []
            for x in eds:
                if t.find(x) != -1: indx.append(t.find(x))
            indx.sort()
            indx_first = indx.copy()
            if len(indx) != 0:
                if indx[0] != 0:
                    t = t[indx[0]:]
                    indx = []
                    for x in eds:
                        if t.find(x) != -1: indx.append(t.find(x))

                if len(indx) > 1:
                    for x in range(1, len(indx)):
                        data.append(t[:indx[x] - 1].strip())
                        t = t[indx[x]:]
                    data.append(t.strip())
                else:
                    if text.strip() not in data:
                        data.append(text.strip())
        else:
            _split_parametr = ', '
            data = list(map(str, text.split(_split_parametr)))
        res = {}
        ed_local_check = [',длина', ',ширина', ',высота', ',диаметр', ',глубина']
        for i in data:
            FLAG = []
            for b in ed_local_check:
                if b in i:
                    FLAG.append(True)
            if len(FLAG) > 1:
                _indx = []
                divided = []
                for j in range(len(ed_local_check)):
                    if i.find(eds[j]) != -1: _indx.append(i.find(eds[j]))
                _indx.sort()
                data.remove(i)
                t = i
                for x in range(1, len(_indx)):
                    data.append(t[:_indx[x] - 1].strip())
                    t = t[_indx[x]:]
                data.append(t.strip())
        for i in range(len(data)):
            last = data[i]
            _chr = last[len(last) - 1:]
            if _chr == ',':
                data[i] = last[:len(last) - 1]
        ed_local_check = [',длина', ',ширина', ',высота', ',диаметр', ',глубина']
        for i in data:
            FLAG = []
            for b in ed_local_check:
                if b in i:
                    FLAG.append(True)
            if len(FLAG) >= 1:
                _indx = []
                divided = []
                for j in range(len(ed_local_check)):
                    if i.find(eds[j]) != -1: _indx.append(i.find(eds[j]))
                _indx.sort()
                data.remove(i)
                t = i
                for x in range(1, len(_indx)):
                    data.append(t[:_indx[x] - 1].strip())
                    t = t[_indx[x]:]
                data.append(t.strip())
        if type == 'all':
            for item in data:
                it = item
                if ' мм' in it:
                    k = 10
                elif ' кг' in it:
                    k = 0.001
                else:
                    k = 1
                # it = it.replace(',', '.')
                # if '-' in it:
                #    it = it[it.find('-')+1:]
                if 'вес' in it and res.setdefault('weight') is None:
                    it = it.replace(',', '.')
                    w = 0
                    if text.lower().count('вес') > 1:
                        for i in data:
                            if 'вес' in i:
                                w += round(float(re.sub("[^0-9,]", "", i).replace(',', '.')) / k, 2)
                    else:
                        w = round(float(re.sub("[^0-9,]", "", it).replace(',', '.')) / k, 2)
                    res['weight'] = w

                for i in range(len(eds)):
                    if eds[i] in it and res.setdefault(eds_en[i]) is None:
                        r = re.sub("[^0-9,-/]", "", it).replace(',', '.').replace('/', '-').lstrip().rstrip()
                        if r == '':
                            res[eds_en[i]] = 'универсальный (растягивается)'
                            break
                        if '-' in r:
                            r_lst = list(map(str, r.split('-')))
                            try:
                                r_lst.remove('')
                            except ValueError:
                                pass
                            av = 0
                            if '.' in r_lst: r_lst.remove('.')
                            if '' in r_lst: r_lst.remove('')
                            if ' ' in r_lst: r_lst.remove(' ')
                            for ind in range(len(r_lst)):
                                if r_lst[ind][:1] == '.':
                                    r_lst[ind] = r_lst[ind][1:]

                            # print(r_lst)
                            for j in r_lst:
                                av += float(j)
                            r = round(float(av / len(r_lst)) / k, 2)
                        else:
                            if r == '.':
                                r = 'универсальный (растягивается)'
                            else:
                                # print([r,data])
                                if r[:1] == '.':
                                    delete = r.count('.') - 1
                                    r = r[delete:]
                                r = round(float(r) / k, 2)
                        res[eds_en[i]] = r
                        # res.setdefault(eds_en[i],r)
                        break
            return res
        if type == 'clothe':
            SIZES = {1: 'XS', 2: 'S', 3: 'M', 4: 'L', 5: 'XL', 6: 'XXL', 7: 'XXXL', 8: '4XL'}
            SIZES_RU = {'XS': '40-42', 'S': '42-44', 'M': '44-46', 'L': '46-48', 'XL': '48-50', '1XL': '52-54',
                        'XXL': '54-56',
                        'XXXL': '56-58', '4XL': '58-60', 'XL': '50-52'}
            eds = ['об. талии', 'об. груди', 'об. бедер']
            indx_obhvats = []
            for i in eds:
                if i in text:
                    indx_obhvats.append(text.index(i))
            if len(data) != 0:
                eds = ['длина', 'ширина', 'высота', 'диаметр', 'глубина']
                eds_en = ['length', 'width', 'height', 'diameter', 'depth']
                for x in data:
                    temp = x
                    if ' мм' in temp:
                        k = 10
                    elif ' кг' in temp:
                        k = 0.001
                    else:
                        k = 1
                    for z in range(len(eds)):
                        if eds[z] in x:
                            t = re.sub("[^0-9,-]", "", x).replace(',', '.')
                            if res.setdefault(eds_en[z]) is None:
                                res[eds_en[z]] = t
                    if 'вес' in temp and res.setdefault('weight') is None:
                        it = temp.replace(',', '.')
                        w = 0
                        if text.lower().count('вес') > 1:
                            for i in data:
                                if 'вес' in i:
                                    w += round(float(re.sub("[^0-9,]", "", i)) / k, 2)
                        else:
                            w = round(float(re.sub("[^0-9,]", "", it)) / k, 2)
                        res['weight'] = w

            for k, v in SIZES_RU.items():
                if v in text:
                    res['size'] = k
                    res['size-ru'] = v
            if len(indx_obhvats) != 0 and len(indx_first) != 0:
                if indx_first[len(indx_first) - 1] > indx_obhvats[len(indx_obhvats) - 1]:
                    text = text[:indx_first[0]]
                else:
                    pass
            check = list(map(str, text.split(', ')))

            for i in check:
                if 'об. ' in i and res.setdefault('size') is None:
                    n = re.sub("[^0-9-]", "", i)
                    l = list(map(int, n.split('-')))
                    sr_arf = round((l[1] + l[0]) / 2)
                    if 'груд' in i:
                        SIZE = Functions.for_sizes_parse([99, 103, 107, 111, 115, 119, 123, 127])
                    if 'бедер' in i:
                        SIZE = Functions.for_sizes_parse([103, 107, 111, 115, 119, 123, 127, 131])
                    if 'тали' in i:
                        SIZE = Functions.for_sizes_parse([86, 90, 94, 98, 102, 107, 112, 117])
                    if sr_arf < SIZE['XS']:
                        res['size'] = 'XS'
                        res['size-ru'] = SIZES_RU['S']
                        return res
                    prev_sz = SIZE['XS']
                    for k, v in SIZE.items():
                        if k == 'XS': pass
                        if sr_arf < v and sr_arf > prev_sz:
                            res['size'] = k
                            res['size-ru'] = SIZES_RU[k]
                            return res
                        prev_sz = v

            t = check[0]
            if t.count('.') > 1:
                res['size'] = 'M'
                res['size-ru'] = SIZES_RU['M']
                return res
            if 'универсальный' in t:
                res['size'] = 'M'
                res['size-ru'] = SIZES_RU['M']
                return res
            if '-' in t and len(t) == 5:
                res['size-ru'] = t
                size = t.split('-')[0]
                for k, v in SIZES_RU.items():
                    if size in v:
                        res['size'] = k
                return res
            for k, v in SIZES.items():
                if (v in t or str(k) in t) and res.setdefault('size') is None:
                    res['size'] = v
                    res['size-ru'] = SIZES_RU[v]
                    return res
            return res
        if type == 'cosmetic':
            ed = ''
            if ' мл' in text:
                ed = 'мл'
            if ' г' in text:
                ed = 'г'
            volume = re.sub("[^0-9 *хx,]", "", text)
            volume = volume.replace('  ', ' *', 1).replace('(', '').replace(')', '')
            res.setdefault('volume', volume.strip())
            res.setdefault('ed', ed.strip())
            return res
        return text

    def download_from_google(self, type, path=os.getcwd()):
        success = False
        while not success:
            url_blacklist_items_wb = [
                'https://drive.google.com/file/d/1bMMyC75qNwpHD61CfaM7gU2zT-dR012t/view?usp=sharing',
                'blacklist_brands_wb']
            url_problem_items_wb_id = [
                'https://drive.google.com/file/d/1dE7VrYH_CDfKFTofDzhjbc8Cp7zZ8ios/view?usp=sharing',
                'problem_items_wb_id']
            url_wb_cats = ['https://drive.google.com/file/d/1zqL6RS35CqQaeZMMAnKHFugmImbJi4t8/view?usp=sharing',
                           'wb_cats']
            url_wb_xlsx_1277 = ['https://drive.google.com/file/d/1SVeUg1-AWWZyTgg9RGkwitOSIDT4Ixul/view?usp=sharing',
                                f'wb_1277.xlsx']
            url_wb_xlsx_1299 = ['https://docs.google.com/file/d/163cgrAFCKd01CGG1FhhT70ibF8d9F7B3/view?usp=sharing',
                                'wb_1299.xlsx']

            download_available = [url_blacklist_items_wb, url_problem_items_wb_id, url_wb_cats, url_wb_xlsx_1277]

            google_ids = ['1bMMyC75qNwpHD61CfaM7gU2zT-dR012t', '1dE7VrYH_CDfKFTofDzhjbc8Cp7zZ8ios',
                          '1zqL6RS35CqQaeZMMAnKHFugmImbJi4t8']
            google_names = ['blacklist_brands_wb.txt', 'problem_items_wb_id.txt', 'wb_cats.txt']

            if self.seller_code == '1277':
                google_ids.append('1SVeUg1-AWWZyTgg9RGkwitOSIDT4Ixul')
                google_names.append('wb_1277.xlsx')
            elif self.seller_code == '1299':
                google_ids.append('163cgrAFCKd01CGG1FhhT70ibF8d9F7B3')
                google_names.append('wb_1299.xlsx')

            path2 = Path(path, 'pool', 'SexOptovik', 'google_downloaded')
            try:
                shutil.rmtree(path2)
                os.rmdir(path2)
            except OSError:
                os.mkdir(path2)
            try:
                if type == 0:
                    Functions.google_driver(google_ids, google_names, './pool/SexOptovik/google_downloaded')
                else:
                    Functions.google_driver([google_ids[type]], [google_names[type]],
                                            './pool/SexOptovik/google_downloaded')
                success = True
            except google.auth.exceptions.RefreshError:
                print('Токен устарел. Необходимо заново авторизироваться в аккаунт.')
                if os.path.isfile('./token_drive_v3.pickle'):
                    os.remove('./token_drive_v3.pickle')
                    print('Устаревший токен успешно удален. Необходимо пройти авторищацию заново.')
                    time.sleep(3)
                else:
                    print('Файл токена не найден в текущем местоположении. Выберите его самостоятельно')
                    path_token = Functions.getFolderFile(0, item=' файл токена google')
                    os.remove(path_token)

    @staticmethod
    def init_category(data, cats_wb, extra):
        CONST_COUNT_WORDS = 3
        data_lower = data.lower()
        getting_in_cat_wb = []
        #getting_in_cat_wb = []
        cats_wb = set()
        extra = extra + '. ' + data_lower
        info = list(map(lambda data: data.strip(), data_lower.lower().split('#')))
        info = list(map(lambda info: info.split('>'), info))
        with open('./pool/SexOptovik/google_downloaded/wb_cats.txt') as f:
            for i in f:
                cats_wb.add(i.rstrip().lower())
        cats_wb.remove('')

        for i in info:
            saved = []
            for j in i:
                if 'белье' not in j:
                    for n in j.split():
                        if n in cats_wb:
                            #getting_in_cat_wb.add(n)
                            getting_in_cat_wb.append(n)
                    t = list(map(lambda j: j + ' ', j.split()))
                    for p in saved:
                        t.append(p + ' ')
                    for n in range(CONST_COUNT_WORDS):
                        t.append('')
                    if 'эротик' not in t:
                        t.append('эротик ')
                    all_var = set(map(''.join, product(t, repeat=CONST_COUNT_WORDS)))
                    all_var = list(map(lambda i: i.rstrip().replace(',', ''), all_var))
                    # all_var = list(map(lambda i: i, all_var))
                    for var in all_var:
                        if var in cats_wb:
                            #getting_in_cat_wb.add(var)
                            getting_in_cat_wb.append(var)
                    saved = j.split().copy()
                else:
                    t = data_lower.lower()
                    ed = ['комплекты', 'бдсм', 'топы', 'лифы', 'бюстье', 'стрэпы', 'корсаж', ' маски', 'комбинезон', 'парик']
                    ed_cats = ['комплекты эротик', 'комплекты бдсм', 'топы эротик', 'бюстгальтеры эротик',
                               'бюстгальтеры эротик', 'стрэпы эротик', 'корсеты эротик', 'маски эротик', 'комбинезоны эротик', 'головные уборы эротик']
                    for o in range(len(ed)):
                        if ed[o] in t:
                            #getting_in_cat_wb.add(ed_cats[o])
                            getting_in_cat_wb.append(ed_cats[o])
                    if 'трусики' in t or 'трусы' in t or 'стринги' in t or 'шортики' in t or 'шорты' in t:
                        #getting_in_cat_wb.add('трусы эротик')
                        getting_in_cat_wb.append('трусы эротик')
        if 'вибро' in data_lower:
            ed = ['зажим', 'яйца', 'трус', 'пуля']
            ed_cats = ['зажимы для сосков', 'виброяйца', 'вибротрусики', 'вибропули']
            for o in range(len(ed)):
                if ed[o] in data_lower:
                    #getting_in_cat_wb.add(ed_cats[o])
                    getting_in_cat_wb.append(ed_cats[o])
        ed = ['лубрикант', 'стек ', 'смазки для секс-игрушек', 'наборы и аксессуары', 'клиторальные стимуляторы',
              'костюмы', 'бдсм товары и фетиш > аксессуары', 'ошейник', 'простынь', 'наручник', 'кольцо эрекционное',
              'стимулирующие влагалище',
              'танго', 'стреп', 'стрэп', 'чулки', 'пояс', 'мастурбатор', 'пестисы', 'пестис',
              'массажное масло', 'гель-смазка', 'смазка', 'массажная свеча', 'массажные свечи', 'насадка на пенис',
              'насадка на вибратор', 'насадка на мастурбатор', 'спрей', 'гель', 'крем', 'презерватив',
              'презервативы', 'фаллоимитатор', 'для ванны', 'скраб', 'масло', 'фаллос',
              'расширитель', 'феромонами', 'кольцо', 'маска', 'помпа', 'вакуумный', 'бальзам',
              'вибромассажер', 'батарейки', 'гель для рук', 'клиторальная']

        ed_cats = ['лубриканты', 'стэки эротик', 'уходовые средства эротик', 'наборы игрушек для взрослых', 'вибраторы',
                   'ролевые костюмы эротик', 'комплекты бдсм', 'ошейники эротик', 'простыни бдсм', 'наручники эротик', 'эрекционные кольца',
                   'вагинальные тренажеры', 'трусы эротик', 'стрэпы эротик', 'стрэпы эротик', 'чулки эротик',
                   'пояса эротик',
                   'мастурбаторы мужские', 'пэстис эротик', 'пэстис эротик', 'массажные средства эротик',
                   'лубриканты', 'лубриканты', 'свечи эротик', 'свечи эротик', 'насадки на член',
                   'насадки для вибраторов', 'насадки на мастурбатор', 'лубриканты', 'лубриканты',
                   'уходовые средства эротик', 'презервативы', 'презервативы', 'фаллоимитаторы',
                   'уходовые средства эротив', 'уходовые средства эротик', 'массажные средства эротик',
                   'фаллоимитаторы', 'расширители гинекологические', 'средства с феромонами',
                   'эрекционные кольца', 'маски эротик', 'вакуумные помпы эротик',
                   'вакуумно-волновые стимуляторы', 'уходовые средства эротик', 'вибратор', 'аксессуары для игрушек эротик',
                   'уходовые средства эротик', 'электростимуляторы']

        if 'насадка' in extra:
            if 'фаллоимитатора' in extra\
                    or 'страпона' in extra:
                getting_in_cat_wb.append('насадки на страпон')
            if 'мастурбатор' in extra\
                    or 'мастурбатора' in extra:
                getting_in_cat_wb.append('насадки для мастурбатора')
            if 'член' in extra\
                    or 'члена' in extra:
                getting_in_cat_wb.append('насадки на член')
        for o in range(len(ed)):
            if ed[o] in extra:
                getting_in_cat_wb.append(ed_cats[o])
        res = []
        if len(getting_in_cat_wb) > 1 and 'комплекты бдсм' in getting_in_cat_wb:
            getting_in_cat_wb.remove('комплекты бдсм')
        for i in getting_in_cat_wb:
            cat = i[:1].upper() + i[1:]
            if 'бдсм' in i:
                cat = cat.replace('бдсм', 'БДСМ')
            res.append(cat)
        if len(res) == 0:
            if 'препарат' in extra: res.append('Возбуждающие препараты')
            else: res.append('Наборы игрушек для взрослых')
        elif len(res) > 1 and 'Набор игрушек для взрослых' in res:
            res.remove('Набор игрушек для взрослых')
        return res[0], data.replace(' #', '.')

    def upload_cats(self):
        status_file = os.path.isfile('./pool/SexOptovik/google_downloaded/wb_cats.txt')
        if status_file:
            set_wb_cats = set()
            with open('./pool/SexOptovik/google_downloaded/wb_cats.txt') as f:
                for i in f:
                    set_wb_cats.add(i.lower().rstrip())
            set_wb_cats.remove('')
            return set_wb_cats
        else:
            print('Файл с категориями отсутсвует. Хотите выбрать его вручную?')
            opt = int(input('1 = да\n2 = нет'))
            if opt == 1:
                data_set_wb_cats = Functions.get_wb_cats(self)
            else:
                print('Далее необходимы категории. Либо загрузите из гугл диска, этапом ранее. Либо скачайте '
                      'самостоятельно и выберите.\nЗакрываю программу')
                time.sleep(5)
                sys.exit(0)

    def start(self):
        ALL_MEDIA = {}
        countries = ['Англия', 'США', 'Беларусь', 'Бельгия', 'Бразилия', 'Португалия', 'Германия', 'Голландия',
                     'Гонконг', 'Дания', 'Индия', 'Испания', 'Турция', 'Италия', 'Казахстан', 'Канада', 'Корея',
                     'Латвия', 'Малайзия', 'Нидерланды', 'Норвегия', 'Польша', 'Россия', 'Сенегал', 'Сингапур',
                     'Мексика', 'Тайланд', 'Тибет', 'Тибет', 'Украина', 'Франция', 'Швейцария', 'Швеция', 'Шотландия',
                     'Эстония', 'Япония', 'Китай']

        # data_set_wb_cats = Functions.get_wb_cats(self) #загрузка из выборочного файла
        amount_of_items_in_file = int(input(f'Введите число товаров в готовых файлах XLSX\n'
                                            f'По стандарту - {self.CONST_AMOUNT_OF_XLSX_ITEMS}\nВвод: '))
        self.CONST_AMOUNT_OF_XLSX_ITEMS = amount_of_items_in_file

        # загрузить файлы txt

        choose = -1
        while choose < 0 or choose > 2:
            choose = int(input(f'Загрузить новые данные ?\n 1 - Да\n0 - Нет\n2  -  Только обновить wb_{self.seller_code}.xslx\n'))
        if choose == 1:
            cwd = self.cwd
            path = Path(cwd, 'pool', 'SexOptovik')
            try:
                shutil.rmtree(path)
                os.rmdir(path)
            except OSError:
                os.mkdir(path)
            except FileExistsError:
                print(f'Пожалуйста, закройте файлы из папки {path}')
            except BaseException:
                print('Ошибка. Попробуйте заново')
                sys.exit(0)
            try:
                self.download_from_google(type=0)
            except google.auth.exceptions.TransportError:
                print('Проверьте соединение\nХотите выбрать файл с категориями вручную?\n1 = да\n2 = нет')
                ch = int(input())
                while ch > 2 or ch < 1:
                    ch = int(input('Введите число 1 ил 2'))
                if ch == 1:
                    print('Еще не сделал')
                    sys.exit(0)
                else:
                    sys.exit(0)
            url = 'http://www.sexoptovik.ru/files/all_prod_info.csv'
            file_path = self.download_universal(url, path_def='./SexOptovik')

            url = 'http://www.sexoptovik.ru/files/all_prod_d33_.csv'
            file_path = self.download_universal(url, path_def='./SexOptovik')
        elif choose == 2:
            try:
                path = f'./pool/SexOptovik/google_downloaded/wb_{self.seller_code}.xlsx'
                os.remove(path)
            except OSError:
                google_id = []
                google_name = []
                if self.seller_code == '1277':
                    google_id.append('1SVeUg1-AWWZyTgg9RGkwitOSIDT4Ixul')
                    google_name.append('wb_1277.xlsx')
                elif self.seller_code == '1299':
                    google_id.append('163cgrAFCKd01CGG1FhhT70ibF8d9F7B3')
                    google_name.append('wb_1299.xlsx')
                self.google_driver(google_ids=google_id,file_names = google_name,
                                   path_os_type='./pool/SexOptovik/google_downloaded')

        else:
            print('Продолжаю со старыми данными\n')
            time.sleep(1.5)
        # !!!!!!!!!!!!!!!!!!!!!!!!!!

        # !!!!!!!!!!!!!!!!!!!!!!!!!!

        PATH_GOOGLE_XLSX = f'./pool/SexOptovik/google_downloaded/wb_{self.seller_code}.xlsx'

        # !!!!!!!!!!!!!!!!!!!!!!!!!!

        # needed_file_data = Functions.getFolderFile(0, ' существующих товары')
        # if not needed_file_data:
        #    print('Ошибка файла. Попробуйте заново')
        #    sys.exit(0)

        print('Считаю новые товары...')
        set_of_data_artics, errors_set_data_artics = Functions.getData(self, PATH_GOOGLE_XLSX,
                                                                       seller_code=self.seller_code)
        blacklist_brands = Functions.uploadFromFile(self,
                                                    file_path='./pool/SexOptovik/google_downloaded/blacklist_brands_wb.txt',
                                                    isSet=True)
        blacklist_brands = list(map(lambda item: item.rstrip().lower(), blacklist_brands))

        PROBLEM_ITEMS = Functions.uploadFromFile(self,
                                                 file_path='./pool/SexOptovik/google_downloaded/problem_items_wb_id.txt',
                                                 isSet=True)
        PROBLEM_ITEMS = list(map(lambda item: item.rstrip().lower(), PROBLEM_ITEMS))
        abs_new_items = 0
        self.current_cats_wb = self.upload_cats()

        opisanie = Functions.uploadFromFile(self, file_path='./SexOptovik/all_prod_d33_.csv', isSet=False)
        print(f'Ошибки: {errors_set_data_artics}')

        file_path = './SexOptovik/all_prod_info.csv'

        path_100 = f'./!parsed_items_{self.CONST_AMOUNT_OF_XLSX_ITEMS}'
        try:
            os.mkdir(path_100)
        except FileExistsError:
            pass
        path_100 += f'/{self.seller_code}'
        success = False
        while not success:
            try:
                shutil.rmtree(path_100)
                os.rmdir(path_100)
            except FileNotFoundError:
                os.mkdir(path_100)
                success = True
            except PermissionError:
                input('Пожалуйста, закройте все открытые файлы из папки на нажмите любую клавишу.\n')

        success = False
        _path = './!parsed_full'
        while not success:
            try:
                shutil.rmtree(_path)
                os.rmdir(_path)
            except FileNotFoundError:
                os.mkdir(_path)
                _path += f'/{self.seller_code}'
                os.mkdir(_path)
                success = True
            except PermissionError:
                input('Пожалуйста, закройте все открытые файлы из папки на нажмите любую клавишу.\n')
        count_items_100 = 0
        with open(file_path) as file:
            ERRORS_ITEMS_BANNED_BRANDS = set()
            for line in file:
                line = line.replace('&quot;', '')
                DATA = list(map(lambda line: line.replace('"', ''), line.split(';')))
                current_articul_wb_pattern = {}
                # отсев существующих артикулов
                if DATA[0] not in set_of_data_artics:
                    if DATA[4].lower() not in blacklist_brands:
                        curr_row_data = {''}
                        abs_new_items += 1
                        # ---                ЗАПОЛНЕНИЕ ШАБЛОНА WB
                        articul = f'id-{DATA[0]}-{self.seller_code}'
                        if articul not in PROBLEM_ITEMS:
                            # id-18474-1277
                            # if count_items_100 == 5179:
                            #     input()
                            model = name = description = category = extra_info = brand = country = osobennost_model = \
                                vibro = elite = sex = colour = complect = count_items = sostav = photo_str = \
                                material = batteries = opis = check = volume = ed = \
                                length_it = length_up = width_it = width_up = weight = weight_bez_up_kg = \
                                diameter_it = diameter_up = depth_it = depth_up = height_it = height_up = photo_wb = \
                                check_brand = check_photo = h = ''

                            h = opisanie[DATA[0]].replace('"', '').rstrip()
                            opis = f'{h}. '
                            model = DATA[1]  # Модель
                            name = Functions.cleanText(DATA[2]).replace('&amp', '')  # Название
                            if name == '':  name = list(map(''.join, opisanie[DATA[0]].split('.; ')))[1].replace('"',
                                                                                                                 '')
                            opis += f'{name}. Модель: {model}. '
                            description = DATA[6]  # Особенность секс игрушки
                            opis += f'{description}. '
                            category_init_data = self.init_category(DATA[3], self.current_cats_wb,
                                                                    extra=f'{name.lower()}')  # Категория
                            category = category_init_data[0]  # Категория
                            extra_info = category_init_data[1]  # Доп. описание (смежные категории)
                            opis += f'{extra_info}. '
                            brand = DATA[4]  # Бренд
                            country = DATA[5]  # Страна производства
                            if '-' in country:
                                country = country[:country.find('-')]
                            osobennost_model = DATA[7][:1].upper() + DATA[7][1:]  # Особенность модели
                            # + Особенность модели
                            if 'с вибрацией' in extra_info:
                                opis += 'С вибрацией. '
                            elif 'без вибрации' in extra_info:
                                opis += 'Без вибрации. '

                            if 'Элитная продукция' in extra_info:
                                opis += 'Элитная продукция. '

                            sex = DATA[8]  # Пол
                            colour = DATA[9]  # Цвет
                            if colour.strip() != '':
                                opis += 'Цвет: ' + colour + '. '
                            count_items = ''
                            if DATA[11] == '':  # Комплектация + Количество предметов в упаковке
                                complect = count_items = '1 шт.'
                            else:
                                complect = DATA[11]
                                count_items = '1 шт.'
                            if DATA[12] != '':
                                sostav = DATA[12]
                            else:
                                sostav = DATA[2]
                            photo_str = DATA[13]
                            photo_urls = list(map(lambda
                                                      photo_str: f'http://sexoptovik.ru/_project/user_images/prods_res/{DATA[0]}/{DATA[0]}_{photo_str}_{self.size_img}.jpg',
                                                  photo_str.split()))
                            material = DATA[15]  # Материал изделия
                            try:
                                material = int(material) - 3
                                material = ''
                                if material != '':
                                    checked = list(map(str, material))[0]
                                    try:
                                        checked = int(checked)
                                        material = ''
                                    except ValueError:
                                        pass
                            except ValueError:
                                if material.strip() != '':
                                    opis += 'Материал: ' + material + '. '
                            # Наличие батареек в комплекте
                            if 'не входят' in DATA[16] or DATA[16] == '':
                                opis += 'Батареек нет в комплекте. '
                            elif 'входят' in DATA[16]:
                                batteries = 'Батарейки в комплекте. '
                                opis += batteries
                            price = 99999  # Розничная цена

                            try:
                                b = list(map(int, DATA[16]))
                                pass
                            except ValueError:
                                opis += DATA[16]

                            check = DATA[2].lower() + '; ' + DATA[3].lower()
                            clothe = ['трусики' in check, 'трусы' in check, 'боди' in check, ' топ ' in check,
                                      'костюм' in check, ' лиф' in check, 'стреп' in check, 'стрэп' in check,
                                      'трус' in check
                                      ]

                            size_all = size_ru = volume = ''
                            if 'Белье' in description or 'БДСМ' in description or any(clothe):
                                size = self.parse_sizes(DATA[10], 'clothe')
                                size_all = size.get('size')
                                size_ru = size.get('size-ru')
                            elif 'Косметика' in description:
                                size = self.parse_sizes(DATA[10], 'cosmetic')
                                volume = size.get('volume')
                                weight = size.get('weight')  # Вес (г)
                                # if '*' in volume: volume = eval(volume)
                                try:
                                    if weight is None:
                                        weight = volume * 1.1
                                        weight_bez_up_kg = volume
                                except TypeError:
                                    pass
                                ed = size.get('ed')
                            else:
                                size = self.parse_sizes(DATA[10], 'all')

                                length_it = size.get('length')  # Длина
                                if length_it is None:
                                    length_it = length_up = ''
                                elif length_it == 'универсальный (растягивается)':
                                    length_it = length_up = 'универсальный (растягивается)'
                                else:
                                    length_up = round(length_it + length_it * 0.1, 3)

                                weight = size.get('weight')  # Вес (г)
                                if weight is None:
                                    weight = weight_bez_up_kg = ''
                                else:
                                    weight_bez_up_kg = int(weight) / 1000
                                width_it = size.get('width')  # Ширина предмета + ширина упаковки
                                if width_it is None:
                                    width_it = width_up = ''
                                elif width_it == 'универсальный (растягивается)':
                                    width_it = width_up = 'универсальный (растягивается)'
                                else:
                                    width_up = round(width_it + width_it * 0.1, 3)

                                height_it = size.get('height')  # Высота предмета + высота упаковки
                                if height_it is None:
                                    height_it = height_up = ''
                                elif height_it == 'универсальный (растягивается)':
                                    height_it = height_up = 'универсальный (растягивается)'
                                else:
                                    height_up = round(height_it + height_it * 0.1, 3)

                                diameter_it = size.get('diameter')  # Диаметр
                                if diameter_it is None:
                                    diameter_it = diameter_up = ''
                                elif diameter_it == 'универсальный (растягивается)':
                                    diameter_it = diameter_up = 'универсальный (растягивается)'
                                else:
                                    diameter_up = round(diameter_it + diameter_it * 0.1, 3)

                                depth_it = size.get('depth')  # Глубина
                                if depth_it is None:
                                    depth_it = depth_up = ''
                                elif depth_it == 'универсальный (растягивается)':
                                    depth_it = depth_up == 'универсальный (растягивается)'
                                else:
                                    depth_up = round(depth_it + depth_it * 0.1, 3)
                            if len(photo_urls) != 0:
                                for z in range(len(photo_urls) - 1):
                                    photo_wb += photo_urls[z] + '; '
                                photo_wb += photo_urls[len(photo_urls) - 1]

                            #   ПРОВЕРКИ
                            PREDLOGS_RU_LANG = {'c', 'в', 'у', 'о', 'к', 'от', 'до', 'на', 'по', 'со', 'из', 'над',
                                                'под', 'при', 'про', 'без', 'ради', 'близ', 'перед', 'около', 'через',
                                                'вдоль', 'после',
                                                'кроме', 'сквозь', 'вроде', 'вследствие', 'благодаря', 'вопреки',
                                                'согласно',
                                                'навстречу', 'об', '(', ')', 'и', ',', 'за', 'и'
                                                }

                            temp = list(map(''.join, name.split()))
                            while len(name) >= 40:
                                name = name[:name.rfind(' ')]
                            while temp[len(temp) - 1] in PREDLOGS_RU_LANG:
                                name = name[:name.rfind(temp[len(temp)-1])]
                                temp = list(map(''.join, name.split(' ')))
                            # check_name = list(map(str,name.split()))
                            # last = check_name[len(check_name)-1]
                            # if last in 'для с без' and len(last)==1 and not(last.isdigit()):
                            #     check_name.pop(len(check_name)-1)
                            #     name = ''.join(check_name)


                            check_brand = brand.upper()
                            if check_brand != brand or '(' in brand or ')' in brand or brand == '':
                                brand = 'Lasciva'
                            opis += f'Бренд: {brand}. '
                            if '28172' in articul:
                                pass

                            if photo_wb == '':
                                for i in DATA:
                                    try:
                                        if i.strip() == '':
                                            flag = True
                                        else:
                                            flag = False
                                        check_photo = list(map(str, i.split()))
                                        for k in check_photo:
                                            if len(k) >= 1 and len(k) <= 2:
                                                flag = True
                                                break
                                        if flag:
                                            pass
                                        else:
                                            photo_wb = ''
                                            photo_urls = list(map(lambda
                                                                      check_photo: f'http://sexoptovik.ru/_project/user_images/prods_res/{DATA[0]}/{DATA[0]}_{check_photo}_{self.size_img}.jpg',
                                                                  check_photo))
                                            for z in range(len(photo_urls) - 1):
                                                photo_wb += photo_urls[z] + '; '
                                            photo_wb += photo_urls[len(photo_urls) - 1]
                                            break
                                    except ValueError:
                                        pass

                            flag = False
                            for datas in DATA:
                                for i in range(len(countries)):
                                    if countries[i] in datas:
                                        country = countries[i]
                                        flag = True
                                        break
                                if flag:
                                    break

                            #           размеры если надо
                            ed_cats = ['Костюм', 'Комплект', 'Наручники',
                                       'Трусы', 'Трусики', 'Стреп', 'Стрэп',
                                       'Чулки', 'Боди', 'Топ', 'Ошейник', 'Платья эротик']
                            for hj in range(len(ed_cats)):
                                if ed_cats[hj] in category and size_ru == '':
                                    size_ru, size_all = '44-46', 'M'

                            # opis = f'{name}. Модель: {model}. Цвет: {colour} Особенности модели: {osobennost_model}. {description}.' \
                            #       f' Бренд: {brand}. {extra_info}. {for_opis}' \
                            #       f'. Наличие батареек в компекте: {batteries}. Материал: {material}. {DATA[10]}'
                            opis = opis.replace('. .', '. ').replace(';', '.').replace('..', '. ')
                            current_articul_wb_pattern = {2: category, 3: '', 4: brand, 5: sex, 6: name,
                                                          7: articul, 8: size_all, 9: size_ru,
                                                          10: '',
                                                          11: price, 12: sostav, 13: photo_wb, 14: opis, 15: country,
                                                          16: osobennost_model, 17: osobennost_model, 18: material,
                                                          19: batteries, 20: f'{volume} {ed}', 21: volume,
                                                          22: f'{volume} {ed}',
                                                          23: width_it, 24: width_up, 25: length_up, 26: length_it,
                                                          27: length_it, 28: height_it, 29: height_up,
                                                          30: depth_it, 31: depth_up, 32: diameter_up, 33: diameter_it,
                                                          34: elite + '. ' + vibro, 35: weight_bez_up_kg, 36: weight,
                                                          37: weight, 38: weight_bez_up_kg, 39: weight, 40: complect,
                                                          41: count_items, 42: 'Непрозрачная анонимная упаковка'
                                                          }

                            count_items_100 += 1
                            for k, v in current_articul_wb_pattern.items():
                                self.parsed_items[k].append(v)
                                self.parsed_items_100_items[k].append(v)
                            if count_items_100 % self.CONST_AMOUNT_OF_XLSX_ITEMS == 0:
                                for kz in range(1, self.CONST_AMOUNT_OF_XLSX_ITEMS + 1):
                                    self.parsed_items_100_items[1].append(kz)
                                Functions.save_data(self, self.parsed_items_100_items, seller_code=self.seller_code,
                                                    path=path_100,
                                                    original_name=f'{count_items_100 - self.CONST_AMOUNT_OF_XLSX_ITEMS - 1}-{count_items_100}')
                                self.parsed_items_100_items.clear()
                                self.parsed_items_100_items = {1: ['Номер карточки'], 2: ['Категория'], 3: ['Цвет'],
                                                               4: ['Бренд'], 5: ['Пол'], 6: ['Название'],
                                                               7: ['Артикул товара'], 8: ['Размер'], 9: ['Рос. размер'],
                                                               10: ['Баркод товара'],
                                                               11: ['Цена'], 12: ['Состав'], 13: ['Медиафайлы'],
                                                               14: ['Описание'],
                                                               15: ['Страна производства'],
                                                               16: ['Особенности секс игрушки'],
                                                               17: ['Особенности модели'], 18: ['Материал'],
                                                               19: ['Наличие батареек в комплекте'], 20: ['Объем'],
                                                               21: ['Объем (мл)'],
                                                               22: ['Объем средства'], 23: ['Ширина предмета'],
                                                               24: ['Ширина упаковки'], 25: ['Длина (см)'],
                                                               26: ['Длина секс игрушки'],
                                                               27: ['Рабочая длина секс игрушки'],
                                                               28: ['Высота предмета'], 29: ['Высота упаковки'],
                                                               30: ['Глубина предмета'], 31: ['Глубина упаковки'],
                                                               32: ['Диаметр'],
                                                               33: ['Диаметр секс игрушки'], 34: ['Вид вибратора'],
                                                               35: ['Вес без упаковки'], 36: ['Вес(г)'],
                                                               37: ['Вес средства'], 38: ['Вес товара без упаковки(г)'],
                                                               39: ['Вес товара с упаковкой(г)'],
                                                               40: ['Комплектация'],
                                                               41: ['Количество предметов в упаковке'], 42: ['Упаковка']
                                                               }

                            ALL_MEDIA.setdefault(articul, photo_urls)
                            print(f'{abs_new_items}, ---     > {current_articul_wb_pattern}')
                    else:
                        ERRORS_ITEMS_BANNED_BRANDS.add(DATA[0])
            for i in range(1, len(self.parsed_items[2])):
                self.parsed_items[1].append(i)
            for i in range(1, len(self.parsed_items_100_items[2])):
                self.parsed_items_100_items[1].append(i)
            Functions.save_data(self, self.parsed_items, seller_code=self.seller_code,
                                path=f'./!parsed_full/{self.seller_code}', _print=False, _full=True)
            Functions.save_data(self, self.parsed_items_100_items, seller_code=self.seller_code, path=path_100,
                                original_name=f'{count_items_100}-END')
            # COUNT_ITEMS_ALLPRODINFO = abs_new_items
            # K = 10 ** 10
            # while K >= 0:
            #    #K = COUNT_ITEMS_ALLPRODINFO - len(SET_BARCODES)
            #    print(f'Необходимо дозагрузить еще {K} штрихкодов')
            #    BARCODES_PATH = Functions.getFolderFile(0,
            #                                            ' файл со штрихкодами.\nФайл должен находиться не на диске C:/ ')
            #    barcodes = Functions.uploadBarcodes(BARCODES_PATH).copy()
            #    if not barcodes:
            #        print('Ошибка. Пожалуйста, выберите необходимый файл с WildBerries')

            print(f'Новых товаров найдено:   {abs_new_items}\nЗапрещенных брендов: {len(ERRORS_ITEMS_BANNED_BRANDS)}')

        input('Обработка товаров с сайта Sex Optovik завершена.')

    # -------------------------------------------------------------------
    wb_pattern = {1: ['Номер карточки'], 2: ['Категория'], 3: ['Цвет'], 4: ['Бренд'], 5: ['Пол'],
                  6: ['Название'], 7: ['Артикул товара'], 8: ['Размер'],
                  9: ['Рос. размер'], 10: ['Баркод товара'], 11: ['Цена'], 12: ['Состав'], 13: ['Описание'],
                  14: ['Вес (г)'], 15: ['Вес без упаковки (кг)'], 16: ['Вес с упаковкой (кг)'],
                  17: ['Вес товара без упаковки (г)'],
                  18: ['Вес товара с упаковкой (г) (г)'], 19: ['Вид вибратора'], 20: ['Вид лубриканта'],
                  21: ['Вид мастурбатора'],
                  22: ['Вкус'], 23: ['Возрастные ограничения'],
                  24: ['Высота предмета (см)'], 25: ['Высота упаковки (см)'], 26: ['Глубина предмета'],
                  27: ['Глубина упаковки (см)'], 28: ['Диаметр предмета (см)'], 29: ['Действие лубриканта'],
                  30: ['Диаметр секс игрушки'], 31: ['Количество предметов в упаковке'],
                  32: ['Количество предметов в упаковке (шт.)'], 33: ['Комплектация'],
                  34: ['Материал изделия'], 35: ['Модель'],
                  36: ['Объем'], 37: ['Объем средства'], 38: ['Объем товара'], 39: ['Особенности белья'],
                  40: ['Особенности продукта'], 41: ['Особенности секс игрушки'],
                  42: ['Рабочая длина секс игрушки'], 43: ['Питание'],
                  44: ['Упаковка'],
                  45: ['Ширина предмета'], 46: ['Ширина упаковки (см)'], 47: ['Медиафайлы'], 48: ['Страна производства']
                  }

    def __init__(self, seller_code, preview='SexOptovik'):
        self.preview = preview
        Functions.showText('Sex Optovik')
        self.seller_code = seller_code

        # print(f'\n--------------------------------------------'
        #      f'\nSexOptovik\n\ncwd:  {self.cwd}')

        res = self.start
        # if not(res):
        #    print('Чтобы начать заново нажмите Enter\nЧтобы выйти нажмите Escape')

        # p 'Доделать !!!!'



class OzonParser(Functions):
    seller_code = ''

    def __init__(self, seller_code):
        self.seller_code = seller_code

    @staticmethod
    def preparing_for_parsing_files_download():
        path = './pool/SexOptovik/Ozon'
        success = False
        while not success:
            try:
                shutil.rmtree(path)
                os.rmdir(path)
                success = True
            except FileNotFoundError:
                os.mkdir(path)
                success = True
            except OSError:
                input(f'Закройте файлы в папке {path} и нажмите любую клавишу')
        download_id = ['1pp0GDS6TF6WaT8us6NhjYvbBdc3mAQcC', '1SnIBrCVFNwlA2Uti6_gzw54BjlMVeHF5']
        name_ids = ['cats_oz.csv', 'banned_brands.txt']
        if seller_code == 1277:
            download_id.append('1dovcWj7ZivWN635m8LJOrcRGkcUKMkJl')
            name_ids.append('oz-1277.csv')
        Functions.google_driver(google_ids=download_id,
                                file_names=name_ids,
                                path_os_type='./pool/SexOptovik/Ozon')
    @staticmethod
    def initialize_cats():
        cats = []
        path = './pool/SexOptovik/Ozon/cats_oz.csv'
        try:
            with open(path, 'r', encoding='utf-8') as categories_file:
                for line in categories_file:
                    row_data = list(map(str, line.split(';')))
                    cats.append([row_data[0], row_data[1], row_data[2]])
            cats[0][0] = cats[0][0][1:]
            return cats
        except UnicodeDecodeError:
            print('Ошибка кодирования')
            sys.exit(0)

    @staticmethod
    def initialize_extra_info():
        art_inf = {}
        try:
            path = './SexOptovik/all_prod_d33_.csv'
            with open(path, 'r', encoding='utf-8') as extra_info_file:
                for line in extra_info_file:
                    line_array = list(map(str, line.split(';')))
                    art_inf.setdefault(line_array[0], [[Functions.cleanText(line_array[1])], line_array[2]])
            return art_inf
        except UnicodeDecodeError:
            print('Ошибка кодирования файла')
            sys.exit(0)

    @staticmethod
    def initialize_blacklist_of_brands():
        banned_brands = set()
        with open('./pool/SexOptovik/Ozon/banned_brands.txt', 'r', encoding='utf-8') as banned_brands_file:
            # banned_brands.add(banned_brands_file.readline())
            for rows in banned_brands_file:
                banned_brands.add(rows.rstrip())
        return banned_brands

    def start_parsing(self):
        # загрузка всевозможных категорий с Ozon
        self.preparing_for_parsing_files_download()
        cats_arr = self.initialize_cats()
        BANNED_BRANDS = self.initialize_blacklist_of_brands()
        PATH_GOOGLE_XLSX = './pool/SexOptovik/Ozon/oz-1277.csv'
        set_of_data_artics, errors_set_data_artics = Functions.getData(self, PATH_GOOGLE_XLSX,
                                                                       seller_code=self.seller_code, marketplace='oz')
        with open('./SexOptovik/all_prod_info.csv', 'r') as main_data_file:
            COUNT_PARSED_ITEMS = COUNT_PROBLEMS_ITEMS = 0
            for rows in main_data_file:
                DATA = list(map(lambda it: it.replace('"', ''), rows.split(';')))
                art_short = DATA[0]
                if art_short not in set_of_data_artics:
                    if DATA[4].upper() not in BANNED_BRANDS:
                        COUNT_PARSED_ITEMS += 1
                    else:
                        COUNT_PROBLEMS_ITEMS += 1
        print(COUNT_PARSED_ITEMS, COUNT_PROBLEMS_ITEMS, errors_set_data_artics)
        print('Готово !')
        input()

def get_provider():
    print('Введите цифру соответсвующую необходимому вам поставщику')
    for i in range(len(all_providers)):
        print(f'{all_providers[i]} --> {i + 1}')
    print(
        '\nДля того, чтобы загрузить всех поставщиков введите "all" или цифру 0\nЧтобы загрузить товары с сайта http://www.sexoptovik.ru \nВВЕДИТЕ: 4')
    code = input().lower()
    if code == '0' or code == 'all':
        return all_providers
    elif code == '1':
        return [all_providers[0]]
    elif code == '2':
        return [all_providers[1]]
    elif code == '4':
        return [all_providers[3]]
    else:
        return [all_providers[2]]


def setup_providers(choosed_providers):
    r = []
    if 'sex_optovik' in choosed_providers: return 4
    if 'astkol' in choosed_providers: r.append(0)
    if 'kema' in choosed_providers: r.append(1)
    if 'andrey' in choosed_providers: r.append(2)
    return r


if __name__ == '__main__':

    cur_time = time.time()
    if not Path.exists(Path(os.getcwd(), 'pool')):
        os.mkdir(Path(os.getcwd(), 'pool'))

    # START
    seller_code = int(input('Введите код продавца: '))

    #OzonParser = OzonParser(f'{seller_code}')
    #OzonParser.start_parsing()

    # input()

    import SexOptovik_wb
    optovik = SexOptovik_wb.SexOptovik(f'{seller_code}')
    optovik.start()

    # выбор поставщика
    func_code = get_provider()
    print(f'Выбрано ->  {func_code}.')
    codes_prov = setup_providers(func_code)

    input()

    parser = Functions()
    # parser.wildberries_site()

    barc = Functions.getFolderFile(0, ' штрихкоды ')
    barcodes = parser.uploadBarcodes(barc).copy()

    data_path = Functions.getFolderFile(0, ' уже добавленные товары ')
    data_set = Functions.getData(data_path, seller_code='1277')

    kema = kema_parser('1277', barcodes, data_set)
    a = parser.workWithExcel(1)
    all_media_files = kema.start_parser()

    ch = input('Вы хотите загрузить фото-медиа файлы?\nДа/Нет\n').lower()
    if ch == 'да':
        media_path = Functions.downloadMedia(all_media_files)
        print(f'Успешно обновлено {len(all_media_files)} товаров. Фотографии находятся в папке{media_path}')
        input('\nВы загрузили их на сайт?')
    time.sleep(0.5)
    print(f'Готово !\nВремя выполнения: {round(time.time() - cur_time)}')
